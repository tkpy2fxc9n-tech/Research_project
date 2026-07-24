# Logic shared by the 4 methods (simulation, training, rollout, export).
# Each methode_*/main.py only chooses INPUT_FIELDS and calls these
# functions, to guarantee that the methods differ ONLY by their inputs.
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor
from contextlib import contextmanager
import fcntl
import os
import platform
import socket
import subprocess
import time

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.animation as animation

import torch
import torch.nn as nn
from torch.utils.data import DataLoader, TensorDataset

# Threads fixed to 1: training/rollout times are compared across
# the 4 methods, a variable thread count would make them non-comparable.
torch.set_num_threads(1)

FIELD_LABELS = {"U": "u", "Ut": "u_dot", "Uxx": "u_xx"}


@dataclass
class Config:
    E: float = 1
    rho: float = 2
    L: float = 1

    Nt: int = 500
    Nx: int = 100
    SS: int = 10
    t_end: float = 5

    # M_BACK past levels -> N_FWD future horizons, spaced by ndt steps.
    ndt: int = 3
    M_BACK: int = 2
    N_FWD: int = 2

    N_GRID: int = 10
    AMP_MIN: float = 0.005
    AMP_MAX: float = 0.1
    OMEGA_MIN: float = 3
    OMEGA_MAX: float = 10

    # None -> grid center (see __post_init__), not the corner (A_MIN, OMEGA_MIN).
    ROLLOUT_A_IDX: int | None = None
    ROLLOUT_OMEGA_IDX: int | None = None

    HIDDEN_SIZES: tuple = (64, 32, 16)

    LEARNING_RATE: float = 1e-3
    N_EPOCHS: int = 18
    BATCH_SIZE: int = 512

    NOISE_STD: float = 0.10
    SMOOTH_ALPHA: float = 0.20   # must stay < 0.25 (smoothing stability)

    SEED: int = 0
    SPLIT_SEED: int = 42

    def __post_init__(self):
        self.Ntot = self.Nx + 2 * self.SS
        self.i_left = self.SS
        self.i_right = self.Ntot - self.SS
        self.nodes = np.arange(self.i_left, self.i_right)
        self.dt = self.t_end / self.Nt
        self.dx = self.L / (self.Nx - 1)
        self.CFL = self.dt / self.dx * np.sqrt(self.E / self.rho)
        if self.CFL > 1:
            print(f"WARNING: CFL={self.CFL:.3f} > 1 -- the explicit scheme is numerically "
                  f"unstable with these Nt/Nx/t_end/L (the simulation will diverge). Increase Nt "
                  f"and/or reduce Nx to bring CFL back to <= 1.")
        self.AMPLITUDES = np.linspace(self.AMP_MIN, self.AMP_MAX, self.N_GRID).round(3).tolist()
        self.PULSATIONS = np.linspace(self.OMEGA_MIN, self.OMEGA_MAX, self.N_GRID).round(1).tolist()
        if self.ROLLOUT_A_IDX is None:
            self.ROLLOUT_A_IDX = self.N_GRID // 2
        if self.ROLLOUT_OMEGA_IDX is None:
            self.ROLLOUT_OMEGA_IDX = self.N_GRID // 2

def set_seeds(cfg: Config) -> None:
    torch.manual_seed(cfg.SEED)
    np.random.seed(cfg.SEED)


def jlabel(k: int) -> str:
    return "j" if k == 0 else f"j{k:+d}"


def lag_label(lag: int) -> str:
    return "t" if lag == 0 else f"t-{lag}ndt"


def uxx_field(u: np.ndarray, cfg: Config) -> np.ndarray:
    out = np.zeros(cfg.Ntot)
    i_left, i_right = cfg.i_left, cfg.i_right
    out[i_left:i_right+1] = (u[i_left-1:i_right] - 2*u[i_left:i_right+1] + u[i_left+1:i_right+2]) / cfg.dx**2
    return out


def field_value(field_name: str, get_u, m: int, cfg: Config) -> np.ndarray:
    if field_name == "U":
        return get_u(m)
    if field_name == "Ut":
        return (get_u(m) - get_u(m - cfg.ndt)) / (cfg.ndt * cfg.dt)
    if field_name == "Uxx":
        return uxx_field(get_u(m), cfg)
    raise ValueError(f"Unknown input field: {field_name!r}")


def make_feature_columns(input_fields: list[str], cfg: Config) -> list[str]:
    cols = []
    for lag in range(cfg.M_BACK):
        lab = lag_label(lag)
        for k in range(-cfg.SS, cfg.SS + 1):
            for f in input_fields:
                cols.append(f"{FIELD_LABELS[f]}({lab},{jlabel(k)})")
    return cols


def make_output_columns(cfg: Config) -> list[str]:
    return [f"delta_u@{h}ndt" for h in range(1, cfg.N_FWD + 1)]


def build_window(m_list, get_u, input_fields: list[str], cfg: Config) -> np.ndarray:
    # Column order must stay synchronized with make_feature_columns.
    nodes = cfg.nodes
    n_features = cfg.M_BACK * (2*cfg.SS + 1) * len(input_fields)
    X = np.zeros((len(nodes), n_features), dtype=np.float32)
    col = 0
    for m in m_list:
        field_arrays = {f: field_value(f, get_u, m, cfg) for f in input_fields}
        for k in range(-cfg.SS, cfg.SS + 1):
            for f in input_fields:
                X[:, col] = field_arrays[f][nodes + k]
                col += 1
    return X


# Each (A, omega) is simulated independently: parallelizable over a process
# pool (uses the cpus allocated by Slurm).
def _n_workers_from_env() -> int:
    slurm_cpus = os.environ.get("SLURM_CPUS_PER_TASK")
    if slurm_cpus:
        return max(1, int(slurm_cpus))
    return os.cpu_count() or 1


def normalize_array(values: np.ndarray, cols, norm_stats: pd.DataFrame) -> np.ndarray:
    mu = norm_stats.loc[cols, "mean"].values.astype(np.float32)
    sd = norm_stats.loc[cols, "std"].values.astype(np.float32)
    return (values.astype(np.float32) - mu) / sd


def make_dataloaders(df: pd.DataFrame, INPUTS, OUTPUTS, norm_stats: pd.DataFrame, cfg: Config):
    train_mask = df["split"] == "train"
    val_mask = df["split"] == "val"

    X_train = normalize_array(df.loc[train_mask, INPUTS].values, INPUTS, norm_stats)
    y_train = normalize_array(df.loc[train_mask, OUTPUTS].values, OUTPUTS, norm_stats)
    X_val = normalize_array(df.loc[val_mask, INPUTS].values, INPUTS, norm_stats)
    y_val = normalize_array(df.loc[val_mask, OUTPUTS].values, OUTPUTS, norm_stats)

    train_loader = DataLoader(
        TensorDataset(torch.tensor(X_train), torch.tensor(y_train)),
        batch_size=cfg.BATCH_SIZE, shuffle=True,
    )
    return train_loader, X_val, y_val


class Reseau(nn.Module):
    def __init__(self, n_inputs, n_outputs, hidden_sizes):
        super().__init__()
        couches = []
        taille_entree = n_inputs
        for taille in hidden_sizes:
            couches.append(nn.Linear(taille_entree, taille))
            couches.append(nn.GELU())
            taille_entree = taille
        couches.append(nn.Linear(taille_entree, n_outputs))
        self.reseau = nn.Sequential(*couches)

    def forward(self, x):
        return self.reseau(x)


@dataclass
class TrainResult:
    historique_train: list
    historique_val: list
    meilleure_val: float
    train_time_s: float
    n_params: int


def train_model(modele, train_loader, X_val, y_val, mu_in, sd_in, mu_out, sd_out, cfg: Config,
                 model_path: Path, patience: int | None = None) -> TrainResult:
    # patience=None (default) -> always runs the full cfg.N_EPOCHS, exactly
    # the old behavior (every existing caller). patience=N -> stops as soon
    # as N consecutive epochs pass without a new best val loss (the best
    # checkpoint, already saved to model_path on every improvement, is what
    # gets reloaded at the end either way).
    #
    # Plain one-step (teacher-forcing) training only: predicts the N_FWD
    # horizons from a real ground-truth M_BACK window, averages the error
    # (nn.MSELoss's default mean reduction, over the batch AND the horizons),
    # backpropagates, steps -- no autoregressive rollout, no pushforward.
    criterion = nn.MSELoss()
    optimiseur = torch.optim.Adam(modele.parameters(), lr=cfg.LEARNING_RATE)
    scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(optimiseur, mode="min", factor=0.5, patience=10)

    with torch.no_grad():
        modele(torch.zeros(1, X_val.shape[1]))

    historique_train, historique_val = [], []
    meilleure_val = float("inf")
    epochs_sans_amelioration = 0

    t0 = time.perf_counter()
    for epoch in range(1, cfg.N_EPOCHS + 1):
        modele.train()
        perte_train = 0.0

        for X_batch, y_batch in train_loader:
            optimiseur.zero_grad()

            if cfg.NOISE_STD > 0:
                X_in = X_batch + cfg.NOISE_STD * torch.randn_like(X_batch)
            else:
                X_in = X_batch

            prediction = modele(X_in)
            data_loss = criterion(prediction, y_batch)
            data_loss.backward()
            optimiseur.step()

            perte_train += data_loss.item()

        perte_train /= len(train_loader)

        modele.eval()
        with torch.no_grad():
            pred_val = modele(torch.tensor(X_val)).numpy()
        perte_val = ((pred_val - y_val) ** 2).mean()
        scheduler.step(perte_val)

        historique_train.append(perte_train)
        historique_val.append(perte_val)

        print(f"Epoch {epoch:4d}/{cfg.N_EPOCHS}  —  data: {perte_train:.4f}  |  val: {perte_val:.4f}")

        if perte_val < meilleure_val:
            meilleure_val = perte_val
            torch.save(modele.state_dict(), model_path)
            epochs_sans_amelioration = 0
        else:
            epochs_sans_amelioration += 1

        if patience is not None and epochs_sans_amelioration >= patience:
            print(f"Early stopping at epoch {epoch}: val loss hasn't improved for "
                  f"{patience} epochs (best={meilleure_val:.6f}).")
            break

    train_time_s = time.perf_counter() - t0

    modele.load_state_dict(torch.load(model_path, weights_only=True))
    print(f"Best model reloaded — minimum val: {meilleure_val:.6f}")

    n_params = sum(p.numel() for p in modele.parameters())
    return TrainResult(historique_train, historique_val, meilleure_val, train_time_s, n_params)


def plot_training_curve(result: TrainResult, output_dir: Path):
    fig, ax = plt.subplots(figsize=(8, 4))
    ax.plot(result.historique_train, label="Data (train)")
    ax.plot(result.historique_val, label="Data (val)")
    ax.set_xlabel("Epoch"); ax.set_ylabel("Loss")
    ax.set_title("Learning curve (teacher forcing)")
    ax.set_yscale("log"); ax.legend(); ax.grid(True)
    plt.tight_layout()
    plt.savefig(output_dir / "courbe_apprentissage.png", dpi=150, bbox_inches="tight")
    plt.close()


def evaluate_teacher_forcing(modele, df_test: pd.DataFrame, INPUTS, OUTPUTS, norm_stats: pd.DataFrame, output_dir: Path) -> dict:
    X_new = normalize_array(df_test[INPUTS].values, INPUTS, norm_stats)
    y_true_n = normalize_array(df_test[OUTPUTS].values, OUTPUTS, norm_stats)
    y_true = df_test[OUTPUTS].values

    modele.eval()
    with torch.no_grad():
        y_pred_n = modele(torch.tensor(X_new)).numpy()

    mu_out = norm_stats.loc[OUTPUTS, "mean"].values
    sd_out = norm_stats.loc[OUTPUTS, "std"].values
    y_pred = y_pred_n * sd_out + mu_out

    fig, axes = plt.subplots(1, len(OUTPUTS), figsize=(6*len(OUTPUTS), 6), squeeze=False)
    axes = axes.flatten()
    metrics = {}
    for i, (ax, col) in enumerate(zip(axes, OUTPUTS)):
        y_r, y_p = y_true[:, i], y_pred[:, i]
        ax.scatter(y_r, y_p, alpha=0.4, s=8)
        lim = max(abs(y_r).max(), abs(y_p).max())
        ax.plot([-lim, lim], [-lim, lim], "r--", lw=1, label="perfect prediction")
        ax.set_xlabel(f"{col} real (physical)")
        ax.set_ylabel(f"{col} predicted (physical)")

        mse_norm = ((y_pred_n[:, i] - y_true_n[:, i]) ** 2).mean()
        r2 = 1 - mse_norm / y_true_n[:, i].var()

        ax.set_title(f"{col}\nMSE (norm)={mse_norm:.2e}  |  R²={r2:.3f}")
        ax.legend(); ax.grid(True)
        metrics[col] = {"mse_norm": float(mse_norm), "r2": float(r2)}

    fig.suptitle("Test over the full test split of the dataset", fontsize=14)
    plt.tight_layout()
    plt.savefig(output_dir / "test_predictions.png", dpi=150, bbox_inches="tight")
    plt.close()
    return metrics


def _biais_repos(modele, mu_in, sd_in, mu_out, sd_out, cfg: Config):
    # Network output for a zero input, subtracted from the rollout so the
    # resting zone stays at 0.
    Xz = (np.zeros((len(cfg.nodes), len(mu_in)), dtype=np.float32) - mu_in) / sd_in
    with torch.no_grad():
        return (modele(torch.tensor(Xz)).numpy() * sd_out + mu_out)[0]


def l2_rel(pred, true, eps=1e-12):
    return np.linalg.norm(pred - true) / (np.linalg.norm(true) + eps)


def smape(pred, true):
    m = true != 0
    return np.mean(2*np.abs(true[m] - pred[m]) / (np.abs(true[m]) + np.abs(pred[m])))


def compute_errors(rollout: "RolloutResultGeneral", cfg: Config):
    steps = np.arange(2*cfg.ndt, cfg.Nt + 1, cfg.ndt)
    t_axis = steps * cfg.dt
    nodes = cfg.nodes
    U, U_reel = rollout.U, rollout.U_reel
    l2_list = [l2_rel(U[k, nodes], U_reel[k, nodes]) for k in steps]
    linf_list = [np.max(np.abs(U[k, nodes] - U_reel[k, nodes])) for k in steps]
    smape_list = [100.0 * smape(U[k, nodes], U_reel[k, nodes]) for k in steps]
    return t_axis, l2_list, linf_list, smape_list


def plot_rollout_error(t_axis, l2_list, linf_list, output_dir: Path):
    plt.figure(figsize=(9, 5))
    plt.plot(t_axis, l2_list, "o-", ms=3, label="relative L2 error")
    plt.plot(t_axis, linf_list, "s-", ms=3, label="max absolute error (Linf)")
    plt.yscale("log")
    plt.xlabel("t"); plt.ylabel("error"); plt.grid(True, which="both"); plt.legend()
    plt.title("Rollout error over time")
    plt.savefig(output_dir / "erreur_temps.png", dpi=150, bbox_inches="tight")
    plt.close()


def plot_smape(t_axis, smape_list, output_dir: Path):
    plt.figure(figsize=(9, 5))
    plt.plot(t_axis, smape_list, "s-", ms=3, label="sMAPE")
    plt.xlabel("t"); plt.ylabel("error (%)"); plt.grid(True); plt.legend()
    plt.title("Rollout sMAPE over time")
    plt.savefig(output_dir / "smape_temps.png", dpi=150, bbox_inches="tight")
    plt.close()


def make_rollout_animation(rollout: "RolloutResultGeneral", cfg: Config, output_dir: Path):
    U, U_reel = rollout.U, rollout.U_reel
    nodes = cfg.nodes
    x = np.linspace(0, cfg.L, cfg.Nx)
    frames = np.arange(0, cfg.Nt + 1, cfg.ndt)

    fig_anim, (axA, axB) = plt.subplots(2, 1, figsize=(9, 7), sharex=True)

    ligne_reel, = axA.plot([], [], "r", lw=2, label="real")
    ligne_pred, = axA.plot([], [], "b--", lw=2, label="predicted")
    ymax = np.abs(U_reel[:, nodes]).max() * 1.2
    axA.set_xlim(0, cfg.L); axA.set_ylim(-ymax, ymax)
    axA.set_ylabel("u"); axA.legend(loc="upper right"); axA.grid(True)

    ligne_err, = axB.plot([], [], "k", lw=1.5, label="|predicted - real|")
    err_max = max(np.max([np.abs(U[m, nodes] - U_reel[m, nodes]).max() for m in frames]) * 1.2, 1e-9)
    axB.set_xlim(0, cfg.L); axB.set_ylim(0, err_max)
    axB.set_xlabel("x"); axB.set_ylabel("absolute error"); axB.legend(loc="upper right"); axB.grid(True)

    titre = fig_anim.suptitle("")

    def maj(m):
        ligne_reel.set_data(x, U_reel[m, nodes])
        ligne_pred.set_data(x, U[m, nodes])
        ligne_err.set_data(x, np.abs(U[m, nodes] - U_reel[m, nodes]))
        titre.set_text(f"Wave propagation — t = {m*cfg.dt:.3f}  (step {m})")
        return ligne_reel, ligne_pred, ligne_err, titre

    anim = animation.FuncAnimation(fig_anim, maj, frames=frames, interval=50, blit=False)
    anim.save(output_dir / "propagation_onde.gif", writer="pillow", fps=20, dpi=110)
    plt.close(fig_anim)


def plot_utt_uxx(rollout: "RolloutResultGeneral", cfg: Config, output_dir: Path):
    # PDE check: u_tt as a function of u_xx, real then predicted.
    U, U_reel = rollout.U, rollout.U_reel
    i_left, i_right = cfg.i_left, cfg.i_right
    dt, dx, ndt, Nt, Ntot = cfg.dt, cfg.dx, cfg.ndt, cfg.Nt, cfg.Ntot

    ureel_tt = np.zeros((Nt, Ntot)); ureel_xx = np.zeros((Nt, Ntot))
    for n in range(1, Nt):
        u_prev, u_curr, u_next = U_reel[n-1], U_reel[n], U_reel[n+1]
        ureel_tt[n, i_left:i_right+1] = (u_next[i_left:i_right+1] - 2*u_curr[i_left:i_right+1] + u_prev[i_left:i_right+1]) / dt**2
        ureel_xx[n, i_left:i_right+1] = (u_curr[i_left-1:i_right] - 2*u_curr[i_left:i_right+1] + u_curr[i_left+1:i_right+2]) / dx**2

    snaps_reel = sorted({int(np.clip(round(frac * Nt), 1, Nt - 1)) for frac in (0.1, 0.2, 0.4)})
    plt.figure()
    for n in snaps_reel:
        plt.scatter(ureel_xx[n, i_left+1:i_right], ureel_tt[n, i_left+1:i_right], s=10, label=f"n = {n}")
    plt.xlabel("u_xx (real)"); plt.ylabel("u_tt (real)")
    plt.legend(); plt.grid(); plt.xlim(-10, 10); plt.ylim(-10, 10)
    plt.title("u_tt as a function of u_xx (real)")
    plt.savefig(output_dir / "utt_uxx_reel.png", dpi=150, bbox_inches="tight")
    plt.close()

    upred_tt = np.zeros((Nt, Ntot)); upred_xx = np.zeros((Nt, Ntot))
    for n in range(ndt, Nt - ndt + 1, ndt):
        u_prev, u_curr, u_next = U[n-ndt], U[n], U[n+ndt]
        upred_tt[n, i_left:i_right+1] = (u_next[i_left:i_right+1] - 2*u_curr[i_left:i_right+1] + u_prev[i_left:i_right+1]) / (ndt*dt)**2
        upred_xx[n, i_left:i_right+1] = (u_curr[i_left-1:i_right] - 2*u_curr[i_left:i_right+1] + u_curr[i_left+1:i_right+2]) / dx**2

    n_max_pred = ((Nt - ndt) // ndt) * ndt
    snaps_pred = sorted({int(np.clip(round(frac * Nt / ndt) * ndt, ndt, n_max_pred)) for frac in (0.01, 0.2, 0.3)})
    plt.figure()
    for n in snaps_pred:
        plt.scatter(upred_xx[n, i_left+1:i_right], upred_tt[n, i_left+1:i_right], s=10, label=f"n = {n}")
    plt.xlabel("u_xx (predicted)"); plt.ylabel("u_tt (predicted)")
    plt.grid(); plt.xlim(-10, 10); plt.ylim(-10, 10); plt.legend()
    plt.title("u_tt as a function of u_xx (prediction)")
    plt.savefig(output_dir / "utt_uxx_predit.png", dpi=150, bbox_inches="tight")
    plt.close()


def chrono(fonction, n_repeat=15, n_warmup=3):
    for _ in range(n_warmup):
        fonction()
    durations = []
    for _ in range(n_repeat):
        t0 = time.perf_counter()
        fonction()
        durations.append(time.perf_counter() - t0)
    d = np.array(durations)
    return d.mean(), d.std(), float(np.median(d))


@dataclass
class BenchmarkResult:
    fd_time_med: float
    fd_time_std: float
    nn_time_med: float
    nn_time_std: float
    flops_per_call: float
    n_calls: int


def _git_commit() -> str:
    # Best-effort: resume.txt should never fail to write just because git
    # is unavailable (e.g. code copied outside the repo).
    try:
        out = subprocess.run(["git", "rev-parse", "--short", "HEAD"], cwd=Path(__file__).resolve().parent,
                              capture_output=True, text=True, timeout=5)
        if out.returncode == 0:
            dirty = subprocess.run(["git", "diff", "--quiet"], cwd=Path(__file__).resolve().parent).returncode != 0
            return out.stdout.strip() + (" (+ uncommitted changes)" if dirty else "")
    except OSError:
        pass
    return "unknown (not a git repo or git unavailable)"


def _input_fields_from_columns(INPUTS) -> list[str]:
    # INPUTS columns look like "u(t,j-3)" / "u_dot(t-1ndt,j+2)" -- the part
    # before "(" is the FIELD_LABELS value, decoded back to "U"/"Ut"/"Uxx" so
    # resume.txt can name the fields without needing input_fields as a
    # separate argument.
    reverse_labels = {v: k for k, v in FIELD_LABELS.items()}
    seen = []
    for col in INPUTS:
        label = col.split("(")[0]
        if label not in seen:
            seen.append(label)
    return [reverse_labels.get(lab, lab) for lab in seen]


def _write_resume_body(f, cfg: Config, method_name: str, df: pd.DataFrame, INPUTS, OUTPUTS,
                        train_result: TrainResult, tf_metrics: dict, bench: BenchmarkResult,
                        errors, rollout_line: str, extra_info: dict | None = None):
    t_axis, l2_list, linf_list, smape_list = errors
    l2_final, l2_max = l2_list[-1], max(l2_list)
    linf_final, linf_max = linf_list[-1], max(linf_list)
    smape_final, smape_max = smape_list[-1], max(smape_list)

    f.write("=====  RUN SUMMARY  =====\n\n")
    f.write(f"Method          : {method_name}\n\n")

    f.write("--- Environment ---\n")
    f.write(f"Run timestamp   : {datetime.now():%Y-%m-%d %H:%M:%S}\n")
    f.write(f"Git commit      : {_git_commit()}\n")
    job_id = os.environ.get("SLURM_JOB_ID")
    f.write(f"Host / job      : {socket.gethostname()} / "
            f"{'SLURM job ' + job_id if job_id else 'no SLURM job id (local run)'}\n")
    f.write(f"Python / torch  : {platform.python_version()} / {torch.__version__} "
            f"({'cuda' if torch.cuda.is_available() else 'cpu'})\n\n")

    f.write("--- Configuration ---\n")
    f.write(f"Grid            : Nt={cfg.Nt}, Nx={cfg.Nx}, SS={cfg.SS}, ndt={cfg.ndt} "
            f"(dt={cfg.dt:.4g}, dx={cfg.dx:.4g}, CFL={cfg.CFL:.3f})\n")
    f.write(f"Physical        : E={cfg.E}, rho={cfg.rho}, L={cfg.L}, t_end={cfg.t_end}\n")
    f.write(f"M / N           : M_BACK={cfg.M_BACK}, N_FWD={cfg.N_FWD}\n")
    f.write(f"Input fields    : {', '.join(_input_fields_from_columns(INPUTS))}\n")
    f.write(f"Stencil         : {2*cfg.SS+1} points (SS={cfg.SS} each side)\n")
    f.write(f"Sampling grid   : {cfg.N_GRID}x{cfg.N_GRID} sims, "
            f"A in [{cfg.AMP_MIN}, {cfg.AMP_MAX}], omega in [{cfg.OMEGA_MIN}, {cfg.OMEGA_MAX}]\n")
    f.write(f"{rollout_line}\n")
    f.write(f"Smoothing/noise : SMOOTH_ALPHA={cfg.SMOOTH_ALPHA}, NOISE_STD={cfg.NOISE_STD}\n")
    f.write(f"Seeds           : SEED={cfg.SEED}, SPLIT_SEED={cfg.SPLIT_SEED}\n")
    f.write(f"Dataset         : {len(df):,} rows\n")
    f.write(f"Features        : {len(INPUTS)} inputs, {len(OUTPUTS)} output(s)\n")
    for s in ["train", "val", "test"]:
        n = (df["split"] == s).sum()
        f.write(f"  split {s:5s}   : {n:>8,} rows ({100*n/len(df):.1f} %)\n")
    f.write(f"NN parameters   : {train_result.n_params:,}\n")
    if extra_info:
        for k, v in extra_info.items():
            f.write(f"{k:15s} : {v}\n")
    f.write("\n")

    f.write("--- Training ---\n")
    f.write(f"Learning rate   : {cfg.LEARNING_RATE:g}\n")
    f.write(f"Minimum val     : {train_result.meilleure_val:.6e}\n")
    for col, m in tf_metrics.items():
        f.write(f"{col:15s} : MSE (norm) = {m['mse_norm']:.4e} | R2 = {m['r2']:.4f}\n")
    f.write("\n")

    f.write("--- Execution time ---\n")
    n_epochs_run = len(train_result.historique_train)
    epochs_label = (f"{n_epochs_run} epochs" if n_epochs_run == cfg.N_EPOCHS
                    else f"{n_epochs_run}/{cfg.N_EPOCHS} epochs, early stopped")
    f.write(f"Training ({epochs_label})         : {train_result.train_time_s:.3f} s\n")
    f.write(f"Real simulation (FD, median)         : {bench.fd_time_med*1e3:.3f} ms\n")
    f.write(f"Predicted rollout (NN, median)       : {bench.nn_time_med*1e3:.3f} ms  (+/-{bench.nn_time_std*1e3:.3f})\n")
    f.write(f"Speedup FD/NN                        : {bench.fd_time_med/bench.nn_time_med:.2f}\n")
    f.write(f"Network FLOPs (full rollout)          : {bench.flops_per_call:,.0f}\n\n")

    f.write("--- Rollout errors ---\n")
    f.write(f"L2 relative  : final = {l2_final:.4e}  |  max = {l2_max:.4e}\n")
    f.write(f"Linf absolute: final = {linf_final:.4e}  |  max = {linf_max:.4e}\n")
    f.write(f"sMAPE (%)    : final = {smape_final:.3f}  |  max = {smape_max:.3f}\n")


# Concurrent writes possible (methods running in parallel): flock serializes
# the load/modify/save of comparative_table.xlsx across processes.
@contextmanager
def _xlsx_lock(xlsx_path: Path):
    lock_path = xlsx_path.with_suffix(xlsx_path.suffix + ".lock")
    with open(lock_path, "w") as f:
        fcntl.flock(f, fcntl.LOCK_EX)
        try:
            yield
        finally:
            fcntl.flock(f, fcntl.LOCK_UN)


def export_errors_to_xlsx(xlsx_path: Path, method_name: str, t_axis, l2_list):
    import openpyxl
    with _xlsx_lock(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

        header = [cell.value for cell in ws[1]]
        if method_name in header:
            col_err = header.index(method_name) + 1
            col_time = col_err - 1
        else:
            col_time = ws.max_column + 1
            col_err = col_time + 1
            ws.cell(row=1, column=col_time, value="t")
            ws.cell(row=1, column=col_err, value=method_name)

        # Clears existing content before writing (avoids ghost rows from a
        # previous run that produced more points).
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_time, value=None)
            ws.cell(row=row, column=col_err, value=None)

        for i, (t, e) in enumerate(zip(t_axis, l2_list), start=2):
            ws.cell(row=i, column=col_time, value=float(t))
            ws.cell(row=i, column=col_err, value=float(e))

        wb.save(xlsx_path)
    print(f"{len(l2_list)} values written to {xlsx_path} (columns t / {method_name}).")


def export_timings_to_xlsx(xlsx_path: Path, method_name: str, train_time_s: float,
                            rollout_time_med_s: float, rollout_time_std_s: float,
                            fd_time_med_s: float, n_params: int, flops_per_call: float):
    import openpyxl
    with _xlsx_lock(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
        if "Timings" not in wb.sheetnames:
            ws = wb.create_sheet("Timings")
            ws.append(["method", "train_time_s", "rollout_time_median_s", "rollout_time_std_s",
                       "fd_time_median_s", "speedup_fd_over_nn", "n_params", "flops_per_rollout_call", "run_timestamp"])
        else:
            ws = wb["Timings"]

        speedup = fd_time_med_s / rollout_time_med_s if rollout_time_med_s else float("nan")
        row_values = [method_name, train_time_s, rollout_time_med_s, rollout_time_std_s,
                      fd_time_med_s, speedup, n_params, flops_per_call, datetime.now().isoformat(timespec="seconds")]

        target_row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == method_name:
                target_row = r
                break
        if target_row is None:
            target_row = ws.max_row + 1

        for col, val in enumerate(row_values, start=1):
            ws.cell(row=target_row, column=col, value=val)

        wb.save(xlsx_path)
    print(f"Timings written to {xlsx_path} (sheet Timings, method {method_name}).")


# ============================================================================
# Generalized boundary conditions (arbitrary Dirichlet/Neumann, either end,
# many families of prescribed movement) -- everything this project's
# pipeline actually uses (the single-family, right-only-forcing physics this
# module originally shipped with has been removed from this vendored copy,
# since Beam_surrogate_model no longer uses it).
#
# A boundary condition is (bc_type, waveform_family, params):
#   bc_type in {"dirichlet", "neumann"} -- Dirichlet prescribes a displacement,
#     Neumann prescribes a slope/flux (0 = a true free end).
#   waveform_family is a key into BC_WAVEFORMS below.
#   params is a dict of that family's sampled numeric parameters.
# The same waveform library is reused for both bc_type and both sides: a
# Dirichlet value and a Neumann flux are both "just some function of time",
# so one shared library of shapes covers both roles.
# ============================================================================
BCSpec = tuple  # (bc_type: str, waveform_family: str, params: dict)

BC_TYPES = ("dirichlet", "neumann")


def sample_gaussian_params(rng, cfg: Config) -> dict:
    return {
        "A": float(rng.uniform(cfg.AMP_MIN, cfg.AMP_MAX)),
        "omega": float(rng.uniform(cfg.OMEGA_MIN, cfg.OMEGA_MAX)),
    }


def gaussian_value(p: dict, t: float) -> float:
    # A single bump, centered so it has mostly risen to zero by t=0
    # (t0 = 4*sigma puts the peak a few sigma in).
    sigma = np.interp(p["omega"], [1.0, 10.0], [0.15, 0.07])
    t0 = 4.0 * sigma
    return p["A"] * np.exp(-((t - t0) / sigma) ** 2)


def sample_sinusoid_params(rng, cfg: Config) -> dict:
    return {
        "A": float(rng.uniform(cfg.AMP_MIN, cfg.AMP_MAX)),
        "omega": float(rng.uniform(cfg.OMEGA_MIN, cfg.OMEGA_MAX)),
        "phase": float(rng.uniform(0.0, 2 * np.pi)),
    }


def sinusoid_value(p: dict, t: float) -> float:
    # Steady back-and-forth vibration -- unlike the Gaussian, this keeps
    # driving the boundary for the whole simulation, not just a single bump.
    return p["A"] * np.sin(p["omega"] * t + p["phase"])


def sample_step_params(rng, cfg: Config) -> dict:
    return {
        "A": float(rng.uniform(cfg.AMP_MIN, cfg.AMP_MAX)),
        "t_onset": float(rng.uniform(0.0, 0.3 * cfg.t_end)),
    }


def step_value(p: dict, t: float) -> float:
    # A sudden jolt that then holds -- tests the network on a step response,
    # qualitatively different from any smooth pulse/oscillation.
    return p["A"] if t >= p["t_onset"] else 0.0


def sample_ramp_params(rng, cfg: Config) -> dict:
    return {
        "A": float(rng.uniform(cfg.AMP_MIN, cfg.AMP_MAX)),
        "duration": float(rng.uniform(0.1 * cfg.t_end, 0.5 * cfg.t_end)),
    }


def ramp_value(p: dict, t: float) -> float:
    # A gradual, steady push up to amplitude A over `duration`, then holds.
    return p["A"] * min(t / p["duration"], 1.0)


def sample_multitone_params(rng, cfg: Config) -> dict:
    # An unpredictable wiggle: a handful of random sine waves added together,
    # instead of a single clean frequency -- the closest of these families to
    # "arbitrary" movement without being literally unconstrained noise.
    n_tones = int(rng.integers(2, 5))
    return {
        "A": rng.uniform(cfg.AMP_MIN, cfg.AMP_MAX, size=n_tones).tolist(),
        "omega": rng.uniform(cfg.OMEGA_MIN, cfg.OMEGA_MAX, size=n_tones).tolist(),
        "phase": rng.uniform(0.0, 2 * np.pi, size=n_tones).tolist(),
    }


def multitone_value(p: dict, t: float) -> float:
    tones = zip(p["A"], p["omega"], p["phase"])
    return sum(A * np.sin(om * t + ph) for A, om, ph in tones) / len(p["A"])


def sample_rest_params(rng, cfg: Config) -> dict:
    return {}


def rest_value(p: dict, t: float) -> float:
    # Nothing happens -- a baseline case (equivalent to amplitude 0).
    return 0.0


BC_WAVEFORMS = {
    "gaussian": (sample_gaussian_params, gaussian_value),
    "sinusoid": (sample_sinusoid_params, sinusoid_value),
    "step": (sample_step_params, step_value),
    "ramp": (sample_ramp_params, ramp_value),
    "random_multitone": (sample_multitone_params, multitone_value),
    "rest": (sample_rest_params, rest_value),
}


def bc_value(bc: BCSpec, t: float) -> float:
    _, family, params = bc
    _, value_fn = BC_WAVEFORMS[family]
    return value_fn(params, t)


def apply_boundary(u: np.ndarray, side: str, bc_type: str, value: float, cfg: Config) -> np.ndarray:
    # Dirichlet: fill the whole ghost band + boundary node with `value`,
    # exactly like the existing fixed-BC code (u_new[:i_left+1]=.../
    # u_new[i_right:]=...) -- unchanged behavior for that branch.
    #
    # Neumann: mirror the ghost band across the boundary with a linear
    # correction so the central-difference slope estimate at the boundary
    # equals `value` (0 -> true free end, reflects without flipping sign).
    # Deliberately does NOT overwrite the boundary node itself: the interior
    # leapfrog stencil already computes a value there using exactly one ghost
    # neighbor, so as long as that neighbor is correctly mirrored, the
    # boundary node comes out right on its own -- overwriting it too (like
    # Dirichlet does) would double-impose a condition that isn't a prescribed
    # value here, only a prescribed slope.
    i_left, i_right, SS, dx = cfg.i_left, cfg.i_right, cfg.SS, cfg.dx
    if side == "left":
        if bc_type == "dirichlet":
            u[:i_left + 1] = value
        else:
            for k in range(1, SS + 1):
                u[i_left - k] = u[i_left + k] - 2 * k * dx * value
    else:
        if bc_type == "dirichlet":
            u[i_right:] = value
        else:
            for k in range(1, SS):  # SS-1 pure ghost points beyond i_right
                u[i_right + k] = u[i_right - k] + 2 * k * dx * value
    return u


def apply_boundary_conditions(u: np.ndarray, t: float, bc_left: BCSpec, bc_right: BCSpec, cfg: Config) -> np.ndarray:
    apply_boundary(u, "left", bc_left[0], bc_value(bc_left, t), cfg)
    apply_boundary(u, "right", bc_right[0], bc_value(bc_right, t), cfg)
    return u


def run_fd_simulation_general(bc_left: BCSpec, bc_right: BCSpec, cfg: Config) -> np.ndarray:
    # Same leapfrog scheme as the beam PDE everywhere in this project,
    # ghost-fill via the shared apply_boundary_conditions helper so either
    # end can be any (bc_type, family) pair instead of the fixed left=0/
    # right=Gaussian-pulse physics.
    i_left, i_right, Ntot = cfg.i_left, cfg.i_right, cfg.Ntot
    u_storage = np.zeros((cfg.Nt + 1, Ntot))
    u = np.zeros(Ntot)
    u_1 = np.zeros(Ntot)
    for n in range(cfg.Nt):
        t = n * cfg.dt
        u_new = np.zeros(Ntot)
        u_new[i_left:i_right+1] = (
            2.0 * u[i_left:i_right+1] - u_1[i_left:i_right+1]
            + cfg.CFL**2 * (u[i_left-1:i_right] - 2.0*u[i_left:i_right+1] + u[i_left+1:i_right+2])
        )
        apply_boundary_conditions(u_new, t + cfg.dt, bc_left, bc_right, cfg)
        u_1, u = u.copy(), u_new
        u_storage[n+1] = u.copy()
    return u_storage


def _simulate_one_general(args):
    idx, bc_left, bc_right, input_fields, cfg, INPUTS, OUTPUTS = args
    u_storage = run_fd_simulation_general(bc_left, bc_right, cfg)
    nodes = cfg.nodes
    n_list = list(range(cfg.M_BACK*cfg.ndt, cfg.Nt - cfg.N_FWD*cfg.ndt + 1))

    X = np.zeros((len(n_list), len(nodes), len(INPUTS)), dtype=np.float32)
    Y = np.zeros((len(n_list), len(nodes), len(OUTPUTS)), dtype=np.float32)
    for i, n in enumerate(n_list):
        m_list = [n - lag*cfg.ndt for lag in range(cfg.M_BACK)]
        X[i] = build_window(m_list, lambda m: u_storage[m], input_fields, cfg)
        for h in range(1, cfg.N_FWD + 1):
            Y[i, :, h-1] = u_storage[n + h*cfg.ndt, nodes] - u_storage[n, nodes]

    meta = pd.DataFrame({"sim_idx": idx, "n_step": n_list})
    df_sim = pd.concat([
        meta.reset_index(drop=True),
        pd.DataFrame(X.reshape(-1, len(INPUTS)), columns=INPUTS),
        pd.DataFrame(Y.reshape(-1, len(OUTPUTS)), columns=OUTPUTS),
    ], axis=1)
    return idx, u_storage, df_sim


def generate_dataset_general(input_fields: list[str], cfg: Config, bc_pairs: list[tuple[BCSpec, BCSpec]],
                              n_workers: int | None = None):
    # ProcessPoolExecutor over a list of (left_bc, right_bc) pairs (random
    # sample, not a dense grid). FIELDS/df are keyed by simulation index (bc
    # dicts aren't hashable, unlike a plain (A,omega) tuple).
    INPUTS = make_feature_columns(input_fields, cfg)
    OUTPUTS = make_output_columns(cfg)

    tasks = [(idx, left, right, input_fields, cfg, INPUTS, OUTPUTS)
             for idx, (left, right) in enumerate(bc_pairs)]
    n_workers = n_workers or min(len(tasks), _n_workers_from_env())

    FIELDS = {}
    dfs = []
    if n_workers > 1:
        with ProcessPoolExecutor(max_workers=n_workers) as ex:
            for idx, u_storage, df_sim in ex.map(_simulate_one_general, tasks):
                FIELDS[idx] = u_storage
                dfs.append(df_sim)
    else:
        for task in tasks:
            idx, u_storage, df_sim = _simulate_one_general(task)
            FIELDS[idx] = u_storage
            dfs.append(df_sim)

    df = pd.concat(dfs, ignore_index=True)
    return df, FIELDS, INPUTS, OUTPUTS


def _autoregressive_rollout_general(modele, U_reel, input_fields, mu_in, sd_in, mu_out, sd_out,
                                     biais_repos, bc_left: BCSpec, bc_right: BCSpec, cfg: Config) -> np.ndarray:
    history_needed = cfg.M_BACK * cfg.ndt
    U = np.zeros((cfg.Nt + 1, cfg.Ntot))
    for m in range(history_needed + 1):
        U[m] = U_reel[m]

    for n in range(history_needed, cfg.Nt - cfg.N_FWD*cfg.ndt + 1, cfg.N_FWD*cfg.ndt):
        m_list = [n - lag*cfg.ndt for lag in range(cfg.M_BACK)]
        X = (build_window(m_list, lambda m: U[m], input_fields, cfg) - mu_in) / sd_in
        with torch.no_grad():
            sortie = modele(torch.tensor(X)).numpy()
        deltas = sortie * sd_out + mu_out - biais_repos

        for h in range(1, cfg.N_FWD + 1):
            s = n + h*cfg.ndt
            t = s * cfg.dt
            U[s, cfg.nodes] = U[n, cfg.nodes] + deltas[:, h-1]
            apply_boundary_conditions(U[s], t, bc_left, bc_right, cfg)

            if cfg.SMOOTH_ALPHA > 0:
                j0, j1 = cfg.i_left + 1, cfg.i_right
                lap = U[s, j0-1:j1-1] - 2*U[s, j0:j1] + U[s, j0+1:j1+1]
                U[s, j0:j1] += cfg.SMOOTH_ALPHA * lap

    return U


@dataclass
class RolloutResultGeneral:
    U: np.ndarray
    U_reel: np.ndarray
    left_bc: "BCSpec"
    right_bc: "BCSpec"


def run_rollout_general(modele, FIELDS: dict, bc_pairs: list[tuple[BCSpec, BCSpec]], rollout_idx: int,
                         input_fields, norm_stats, INPUTS, OUTPUTS, cfg: Config) -> RolloutResultGeneral:
    left_bc, right_bc = bc_pairs[rollout_idx]
    U_reel = FIELDS[rollout_idx]

    mu_in = norm_stats.loc[INPUTS, "mean"].values.astype(np.float32)
    sd_in = norm_stats.loc[INPUTS, "std"].values.astype(np.float32)
    mu_out = norm_stats.loc[OUTPUTS, "mean"].values.astype(np.float32)
    sd_out = norm_stats.loc[OUTPUTS, "std"].values.astype(np.float32)

    biais_repos = _biais_repos(modele, mu_in, sd_in, mu_out, sd_out, cfg)
    U = _autoregressive_rollout_general(modele, U_reel, input_fields, mu_in, sd_in, mu_out, sd_out,
                                         biais_repos, left_bc, right_bc, cfg)
    return RolloutResultGeneral(U=U, U_reel=U_reel, left_bc=left_bc, right_bc=right_bc)


def bc_describe(bc: BCSpec) -> str:
    bc_type, family, params = bc
    param_str = ", ".join(f"{k}={v}" for k, v in params.items())
    return f"{bc_type}/{family}({param_str})"


def benchmark_inference_general(modele, FIELDS, input_fields, norm_stats, INPUTS, OUTPUTS,
                                 rollout: RolloutResultGeneral, cfg: Config) -> BenchmarkResult:
    left_bc, right_bc, U_reel = rollout.left_bc, rollout.right_bc, rollout.U_reel

    mu_in = norm_stats.loc[INPUTS, "mean"].values.astype(np.float32)
    sd_in = norm_stats.loc[INPUTS, "std"].values.astype(np.float32)
    mu_out = norm_stats.loc[OUTPUTS, "mean"].values.astype(np.float32)
    sd_out = norm_stats.loc[OUTPUTS, "std"].values.astype(np.float32)
    biais_repos = _biais_repos(modele, mu_in, sd_in, mu_out, sd_out, cfg)

    def fd_once():
        return run_fd_simulation_general(left_bc, right_bc, cfg)

    def rollout_once():
        return _autoregressive_rollout_general(modele, U_reel, input_fields, mu_in, sd_in, mu_out, sd_out,
                                                biais_repos, left_bc, right_bc, cfg)

    fd_mean, fd_std, fd_med = chrono(fd_once)
    nn_mean, nn_std, nn_med = chrono(rollout_once)

    n_calls = len(range(cfg.M_BACK*cfg.ndt, cfg.Nt - cfg.N_FWD*cfg.ndt + 1, cfg.N_FWD*cfg.ndt))
    n_features = cfg.M_BACK * (2*cfg.SS + 1) * len(input_fields)

    from torch.utils.flop_counter import FlopCounterMode
    with FlopCounterMode(display=False) as fc:
        modele(torch.zeros((len(cfg.nodes), n_features)))
    flops_per_call = fc.get_total_flops() * n_calls

    print(f"FD (real)   : {fd_med*1e3:7.3f} ms")
    print(f"NN (rollout): {nn_med*1e3:7.3f} ms  (±{nn_std*1e3:.3f})")
    print(f"speedup FD/NN = {fd_med/nn_med:.2f}x   (>1 = the network is faster)")

    return BenchmarkResult(
        fd_time_med=fd_med, fd_time_std=float(fd_std),
        nn_time_med=nn_med, nn_time_std=float(nn_std),
        flops_per_call=flops_per_call, n_calls=n_calls,
    )


def export_resume_general(output_dir: Path, cfg: Config, method_name: str, df: pd.DataFrame, INPUTS, OUTPUTS,
                           train_result: TrainResult, tf_metrics: dict, rollout: RolloutResultGeneral,
                           bench: BenchmarkResult, errors, extra_info: dict | None = None):
    rollout_line = f"Rollout BC      : left={bc_describe(rollout.left_bc)}  right={bc_describe(rollout.right_bc)}"
    with open(output_dir / "resume.txt", "w") as f:
        _write_resume_body(f, cfg, method_name, df, INPUTS, OUTPUTS, train_result, tf_metrics, bench,
                            errors, rollout_line, extra_info)

    print(f"Summary saved: {output_dir / 'resume.txt'}")
