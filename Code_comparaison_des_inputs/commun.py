# Logique commune aux 4 méthodes (simulation, entraînement, rollout, export).
# Chaque methode_*/main.py ne fait que choisir INPUT_FIELDS et appeler ces
# fonctions, pour garantir que les méthodes ne diffèrent QUE par leurs inputs.
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from itertools import product
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor
from contextlib import contextmanager
import fcntl
import os
import time

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.animation as animation

import torch
import torch.nn as nn
from torch.utils.data import DataLoader, TensorDataset

# Threads fixés à 1 : les temps d'entraînement/rollout sont comparés entre
# les 4 méthodes, un thread count variable les rendrait non comparables.
torch.set_num_threads(1)

FIELD_LABELS = {"U": "u", "Ut": "u_dot", "Uxx": "u_xx"}


@dataclass
class Config:
    E: float = 1
    rho: float = 2
    L: float = 1

    Nt: int = 500
    Nx: int = 100
    SS: int = 10
    t_end: float = 5

    # M_BACK niveaux passés -> N_FWD horizons futurs, espacés de ndt pas.
    ndt: int = 3
    M_BACK: int = 2
    N_FWD: int = 2

    N_GRID: int = 10
    AMP_MIN: float = 0.005
    AMP_MAX: float = 0.1
    OMEGA_MIN: float = 3
    OMEGA_MAX: float = 10

    # None -> centre de la grille (cf __post_init__), pas le coin (A_MIN, OMEGA_MIN).
    ROLLOUT_A_IDX: int | None = None
    ROLLOUT_OMEGA_IDX: int | None = None

    HIDDEN_SIZES: tuple = (64, 32, 16)

    LEARNING_RATE: float = 1e-3
    N_EPOCHS: int = 20
    BATCH_SIZE: int = 512

    LAMBDA_PF: float = 1.0
    N_PF_GROUPS: int = 8
    PF_WARMUP: int = 2
    PF_HOPS: int = 3   # sauts autorégressifs enchaînés pendant l'entraînement pushforward

    NOISE_STD: float = 0.10
    SMOOTH_ALPHA: float = 0.20   # doit rester < 0.25 (stabilité du lissage)

    SEED: int = 0
    SPLIT_SEED: int = 42

    def __post_init__(self):
        self.Ntot = self.Nx + 2 * self.SS
        self.i_left = self.SS
        self.i_right = self.Ntot - self.SS
        self.nodes = np.arange(self.i_left, self.i_right)
        self.dt = self.t_end / self.Nt
        self.dx = self.L / (self.Nx - 1)
        self.CFL = self.dt / self.dx * np.sqrt(self.E / self.rho)
        if self.CFL > 1:
            print(f"ATTENTION : CFL={self.CFL:.3f} > 1 -- le schéma explicite est numériquement "
                  f"instable avec ces Nt/Nx/t_end/L (la simulation va diverger). Augmentez Nt "
                  f"et/ou réduisez Nx pour revenir à CFL <= 1.")
        self.AMPLITUDES = np.linspace(self.AMP_MIN, self.AMP_MAX, self.N_GRID).round(3).tolist()
        self.PULSATIONS = np.linspace(self.OMEGA_MIN, self.OMEGA_MAX, self.N_GRID).round(1).tolist()
        if self.ROLLOUT_A_IDX is None:
            self.ROLLOUT_A_IDX = self.N_GRID // 2
        if self.ROLLOUT_OMEGA_IDX is None:
            self.ROLLOUT_OMEGA_IDX = self.N_GRID // 2


def set_seeds(cfg: Config) -> None:
    torch.manual_seed(cfg.SEED)
    np.random.seed(cfg.SEED)


def jlabel(k: int) -> str:
    return "j" if k == 0 else f"j{k:+d}"


def lag_label(lag: int) -> str:
    return "t" if lag == 0 else f"t-{lag}ndt"


def u_right_val(A: float, omega: float, t: float) -> float:
    sigma = np.interp(omega, [1.0, 10.0], [0.15, 0.07])
    t0 = 4.0 * sigma
    return A * np.exp(-((t - t0) / sigma) ** 2)


def run_fd_simulation(A: float, omega: float, cfg: Config) -> np.ndarray:
    # Grille étendue avec SS points fantômes : bord gauche encastré (=0),
    # bord droit imposé (continuité constante avec le déplacement forcé).
    i_left, i_right, Ntot = cfg.i_left, cfg.i_right, cfg.Ntot
    u_storage = np.zeros((cfg.Nt + 1, Ntot))
    u = np.zeros(Ntot)
    u_1 = np.zeros(Ntot)
    for n in range(cfg.Nt):
        t = n * cfg.dt
        u_new = np.zeros(Ntot)
        u_new[i_left:i_right+1] = (
            2.0 * u[i_left:i_right+1] - u_1[i_left:i_right+1]
            + cfg.CFL**2 * (u[i_left-1:i_right] - 2.0*u[i_left:i_right+1] + u[i_left+1:i_right+2])
        )
        u_new[:i_left+1] = 0.0
        u_new[i_right:] = u_right_val(A, omega, t + cfg.dt)
        u_1, u = u.copy(), u_new
        u_storage[n+1] = u.copy()
    return u_storage


def uxx_field(u: np.ndarray, cfg: Config) -> np.ndarray:
    out = np.zeros(cfg.Ntot)
    i_left, i_right = cfg.i_left, cfg.i_right
    out[i_left:i_right+1] = (u[i_left-1:i_right] - 2*u[i_left:i_right+1] + u[i_left+1:i_right+2]) / cfg.dx**2
    return out


def field_value(field_name: str, get_u, m: int, cfg: Config) -> np.ndarray:
    if field_name == "U":
        return get_u(m)
    if field_name == "Ut":
        return (get_u(m) - get_u(m - cfg.ndt)) / (cfg.ndt * cfg.dt)
    if field_name == "Uxx":
        return uxx_field(get_u(m), cfg)
    raise ValueError(f"Champ d'entrée inconnu : {field_name!r}")


def make_feature_columns(input_fields: list[str], cfg: Config) -> list[str]:
    cols = []
    for lag in range(cfg.M_BACK):
        lab = lag_label(lag)
        for k in range(-cfg.SS, cfg.SS + 1):
            for f in input_fields:
                cols.append(f"{FIELD_LABELS[f]}({lab},{jlabel(k)})")
    return cols


def make_output_columns(cfg: Config) -> list[str]:
    return [f"delta_u@{h}ndt" for h in range(1, cfg.N_FWD + 1)]


def build_window(m_list, get_u, input_fields: list[str], cfg: Config) -> np.ndarray:
    # L'ordre des colonnes doit rester synchronisé avec make_feature_columns.
    nodes = cfg.nodes
    n_features = cfg.M_BACK * (2*cfg.SS + 1) * len(input_fields)
    X = np.zeros((len(nodes), n_features), dtype=np.float32)
    col = 0
    for m in m_list:
        field_arrays = {f: field_value(f, get_u, m, cfg) for f in input_fields}
        for k in range(-cfg.SS, cfg.SS + 1):
            for f in input_fields:
                X[:, col] = field_arrays[f][nodes + k]
                col += 1
    return X


# Chaque (A, omega) est simulé indépendamment : parallélisable sur un pool
# de process (utilise les cpus alloués par Slurm).
def _n_workers_from_env() -> int:
    slurm_cpus = os.environ.get("SLURM_CPUS_PER_TASK")
    if slurm_cpus:
        return max(1, int(slurm_cpus))
    return os.cpu_count() or 1


def _simulate_one(args):
    A, omega, input_fields, cfg, INPUTS, OUTPUTS = args
    nodes = cfg.nodes
    u_storage = run_fd_simulation(A, omega, cfg)

    n_list = list(range(cfg.M_BACK*cfg.ndt, cfg.Nt - cfg.N_FWD*cfg.ndt + 1))

    X = np.zeros((len(n_list), len(nodes), len(INPUTS)), dtype=np.float32)
    Y = np.zeros((len(n_list), len(nodes), len(OUTPUTS)), dtype=np.float32)
    for i, n in enumerate(n_list):
        m_list = [n - lag*cfg.ndt for lag in range(cfg.M_BACK)]
        X[i] = build_window(m_list, lambda m: u_storage[m], input_fields, cfg)
        for h in range(1, cfg.N_FWD + 1):
            Y[i, :, h-1] = u_storage[n + h*cfg.ndt, nodes] - u_storage[n, nodes]

    meta = pd.DataFrame({
        "A": A, "omega": omega,
        "n_step": np.repeat(n_list, len(nodes)),
    })
    df_sim = pd.concat([
        meta.reset_index(drop=True),
        pd.DataFrame(X.reshape(-1, len(INPUTS)), columns=INPUTS),
        pd.DataFrame(Y.reshape(-1, len(OUTPUTS)), columns=OUTPUTS),
    ], axis=1)
    return A, omega, u_storage, df_sim


def generate_dataset(input_fields: list[str], cfg: Config, n_workers: int | None = None):
    INPUTS = make_feature_columns(input_fields, cfg)
    OUTPUTS = make_output_columns(cfg)

    grid = list(product(cfg.AMPLITUDES, cfg.PULSATIONS))
    tasks = [(A, omega, input_fields, cfg, INPUTS, OUTPUTS) for A, omega in grid]
    n_workers = n_workers or min(len(grid), _n_workers_from_env())

    FIELDS = {}
    frames = []
    if n_workers > 1:
        with ProcessPoolExecutor(max_workers=n_workers) as ex:
            for A, omega, u_storage, df_sim in ex.map(_simulate_one, tasks):
                FIELDS[(A, omega)] = u_storage
                frames.append(df_sim)
    else:
        for task in tasks:
            A, omega, u_storage, df_sim = _simulate_one(task)
            FIELDS[(A, omega)] = u_storage
            frames.append(df_sim)

    df = pd.concat(frames, ignore_index=True)
    return df, FIELDS, INPUTS, OUTPUTS


# Normalisation appliquée à la volée (pas de colonnes normalisées dans df)
# pour limiter le pic mémoire.
def split_and_normalize(df: pd.DataFrame, INPUTS, OUTPUTS, cfg: Config):
    rng = np.random.default_rng(seed=cfg.SPLIT_SEED)

    n_rows = len(df)
    n_train = int(0.90 * n_rows)
    n_val = int(0.05 * n_rows)
    n_test = n_rows - n_train - n_val

    split_labels = np.array(["train"]*n_train + ["val"]*n_val + ["test"]*n_test)
    rng.shuffle(split_labels)

    df = df.copy()
    df["split"] = split_labels

    print("Distribution du split :")
    for s in ["train", "val", "test"]:
        n = (df["split"] == s).sum()
        print(f"  {s:5s} : {n:>8,} lignes  ({100*n/len(df):.1f} %)")

    train_mask = df["split"] == "train"
    cols = INPUTS + OUTPUTS
    norm_stats = pd.DataFrame({
        "mean": df.loc[train_mask, cols].mean(),
        "std": df.loc[train_mask, cols].std(),
    })
    norm_stats["std"] = norm_stats["std"].replace(0, 1)
    return df, norm_stats


def normalize_array(values: np.ndarray, cols, norm_stats: pd.DataFrame) -> np.ndarray:
    mu = norm_stats.loc[cols, "mean"].values.astype(np.float32)
    sd = norm_stats.loc[cols, "std"].values.astype(np.float32)
    return (values.astype(np.float32) - mu) / sd


def make_dataloaders(df: pd.DataFrame, INPUTS, OUTPUTS, norm_stats: pd.DataFrame, cfg: Config):
    train_mask = df["split"] == "train"
    val_mask = df["split"] == "val"

    X_train = normalize_array(df.loc[train_mask, INPUTS].values, INPUTS, norm_stats)
    y_train = normalize_array(df.loc[train_mask, OUTPUTS].values, OUTPUTS, norm_stats)
    X_val = normalize_array(df.loc[val_mask, INPUTS].values, INPUTS, norm_stats)
    y_val = normalize_array(df.loc[val_mask, OUTPUTS].values, OUTPUTS, norm_stats)

    train_loader = DataLoader(
        TensorDataset(torch.tensor(X_train), torch.tensor(y_train)),
        batch_size=cfg.BATCH_SIZE, shuffle=True,
    )
    return train_loader, X_val, y_val


class Reseau(nn.Module):
    def __init__(self, n_inputs, n_outputs, hidden_sizes):
        super().__init__()
        couches = []
        taille_entree = n_inputs
        for taille in hidden_sizes:
            couches.append(nn.Linear(taille_entree, taille))
            couches.append(nn.GELU())
            taille_entree = taille
        couches.append(nn.Linear(taille_entree, n_outputs))
        self.reseau = nn.Sequential(*couches)

    def forward(self, x):
        return self.reseau(x)


# Pushforward (Brandstetter et al. 2022) : corrige le distribution shift du
# rollout autorégressif en entraînant aussi sur ses propres prédictions
# (gradient détaché), pas seulement sur des entrées propres (teacher forcing).
def make_pf_samples(FIELDS: dict, cfg: Config) -> list[tuple]:
    return [
        (A, omega, n)
        for (A, omega) in FIELDS
        for n in range(cfg.M_BACK*cfg.ndt, cfg.Nt - (cfg.PF_HOPS + 1)*cfg.N_FWD*cfg.ndt + 1)
    ]


def reconstruct(u_curr, n_curr, pred_norm, A, omega, mu_out, sd_out, cfg: Config,
                 biais_repos: np.ndarray | None = None) -> dict:
    deltas = pred_norm * sd_out + mu_out
    if biais_repos is not None:
        deltas = deltas - biais_repos
    nodes = cfg.nodes
    champs = {}
    for h in range(1, cfg.N_FWD + 1):
        s = n_curr + h * cfg.ndt
        u = np.zeros(cfg.Ntot)
        u[nodes] = u_curr[nodes] + deltas[:, h-1]
        u[:cfg.i_left+1] = 0.0
        u[cfg.i_right:] = u_right_val(A, omega, s * cfg.dt)
        if cfg.SMOOTH_ALPHA > 0:
            j0, j1 = cfg.i_left + 1, cfg.i_right
            lap = u[j0-1:j1-1] - 2*u[j0:j1] + u[j0+1:j1+1]
            u[j0:j1] += cfg.SMOOTH_ALPHA * lap
        champs[s] = u
    return champs


def pushforward_loss(modele, FIELDS, PF_SAMPLES, input_fields, mu_in, sd_in, mu_out, sd_out,
                      criterion, cfg: Config, n_groups: int, biais_repos: np.ndarray):
    idxs = np.random.choice(len(PF_SAMPLES), n_groups, replace=False)
    groups = [PF_SAMPLES[i] for i in idxs]
    nN = len(cfg.nodes)

    # PF_HOPS sauts autorégressifs enchaînés (gradient détaché sauf le
    # dernier) : chaque saut prédit à partir de SA PROPRE reconstruction du
    # saut précédent, pas des données réelles -- comme en rollout complet.
    field_at = [(lambda m, U=FIELDS[(A, omega)]: U[m]) for (A, omega, n) in groups]
    n_curr = [n for (A, omega, n) in groups]
    baseline = [FIELDS[(A, omega)][n] for (A, omega, n) in groups]

    pred = None
    for hop in range(cfg.PF_HOPS):
        last = hop == cfg.PF_HOPS - 1
        X = np.concatenate([
            (build_window([n_curr[j] - lag*cfg.ndt for lag in range(cfg.M_BACK)],
                           field_at[j], input_fields, cfg) - mu_in) / sd_in
            for j in range(n_groups)
        ], axis=0)

        if last:
            pred = modele(torch.tensor(X))
            pred_np = pred.detach().numpy()
        else:
            with torch.no_grad():
                pred_np = modele(torch.tensor(X)).numpy()

        new_field_at, new_n_curr, new_baseline = [], [], []
        for j, (A, omega, n) in enumerate(groups):
            U = FIELDS[(A, omega)]
            Up = reconstruct(baseline[j], n_curr[j], pred_np[j*nN:(j+1)*nN], A, omega, mu_out, sd_out, cfg,
                              biais_repos=biais_repos)
            nprime = n_curr[j] + cfg.N_FWD * cfg.ndt
            new_field_at.append(lambda m, U=U, Up=Up: Up[m] if m in Up else U[m])
            new_n_curr.append(nprime)
            new_baseline.append(Up[nprime])
        field_at, n_curr, baseline = new_field_at, new_n_curr, new_baseline

    tgt_list = []
    for j, (A, omega, n) in enumerate(groups):
        U = FIELDS[(A, omega)]
        curr = baseline[j][cfg.nodes]
        tgt = np.stack([U[n_curr[j] + h*cfg.ndt][cfg.nodes] - curr for h in range(1, cfg.N_FWD+1)], axis=1)
        tgt_list.append(((tgt - mu_out) / sd_out).astype(np.float32))

    return criterion(pred, torch.tensor(np.concatenate(tgt_list, axis=0)))


@dataclass
class TrainResult:
    historique_train: list
    historique_val: list
    historique_pf: list
    meilleure_val: float
    train_time_s: float
    n_params: int


def train_model(modele, train_loader, X_val, y_val, FIELDS, PF_SAMPLES, input_fields,
                 mu_in, sd_in, mu_out, sd_out, cfg: Config, model_path: Path) -> TrainResult:
    criterion = nn.MSELoss()
    optimiseur = torch.optim.Adam(modele.parameters(), lr=cfg.LEARNING_RATE)
    scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(optimiseur, mode="min", factor=0.5, patience=10)

    with torch.no_grad():
        modele(torch.zeros(1, X_val.shape[1]))

    historique_train, historique_val, historique_pf = [], [], []
    meilleure_val = float("inf")


    t0 = time.perf_counter()
    for epoch in range(1, cfg.N_EPOCHS + 1):
        lam_pf = cfg.LAMBDA_PF * min(1.0, epoch / cfg.PF_WARMUP)

        biais_repos = _biais_repos(modele, mu_in, sd_in, mu_out, sd_out, cfg)

        modele.train()
        perte_train = 0.0
        perte_pf_total = 0.0

        for X_batch, y_batch in train_loader:
            optimiseur.zero_grad()

            if cfg.NOISE_STD > 0:
                X_in = X_batch + cfg.NOISE_STD * torch.randn_like(X_batch)
            else:
                X_in = X_batch

            prediction = modele(X_in)
            data_loss = criterion(prediction, y_batch)

            if lam_pf > 0:
                pf_loss = pushforward_loss(modele, FIELDS, PF_SAMPLES, input_fields,
                                            mu_in, sd_in, mu_out, sd_out, criterion, cfg, cfg.N_PF_GROUPS,
                                            biais_repos=biais_repos)
            else:
                pf_loss = torch.tensor(0.0)

            total_loss = data_loss + lam_pf * pf_loss
            total_loss.backward()
            optimiseur.step()

            perte_train += data_loss.item()
            perte_pf_total += pf_loss.item()

        perte_train /= len(train_loader)
        perte_pf_total /= len(train_loader)

        modele.eval()
        with torch.no_grad():
            pred_val = modele(torch.tensor(X_val)).numpy()
        perte_val = ((pred_val - y_val) ** 2).mean()
        scheduler.step(perte_val)

        historique_train.append(perte_train)
        historique_val.append(perte_val)
        historique_pf.append(perte_pf_total)

        print(f"Epoch {epoch:4d}/{cfg.N_EPOCHS}  —  "
              f"data: {perte_train:.4f}  |  pushf: {perte_pf_total:.4f}  |  val: {perte_val:.4f}")

        if perte_val < meilleure_val:
            meilleure_val = perte_val
            torch.save(modele.state_dict(), model_path)

    train_time_s = time.perf_counter() - t0

    modele.load_state_dict(torch.load(model_path, weights_only=True))
    print(f"Meilleur modèle rechargé — val minimale : {meilleure_val:.6f}")

    n_params = sum(p.numel() for p in modele.parameters())
    return TrainResult(historique_train, historique_val, historique_pf, meilleure_val, train_time_s, n_params)


def plot_training_curve(result: TrainResult, output_dir: Path):
    fig, ax = plt.subplots(figsize=(8, 4))
    ax.plot(result.historique_train, label="Data (train)")
    ax.plot(result.historique_val, label="Data (val)")
    ax.plot(result.historique_pf, label="Pushforward")
    ax.set_xlabel("Epoch"); ax.set_ylabel("Loss")
    ax.set_title("Courbe d'apprentissage (pushforward)")
    ax.set_yscale("log"); ax.legend(); ax.grid(True)
    plt.tight_layout()
    plt.savefig(output_dir / "courbe_apprentissage.png", dpi=150, bbox_inches="tight")
    plt.close()


def evaluate_teacher_forcing(modele, df_test: pd.DataFrame, INPUTS, OUTPUTS, norm_stats: pd.DataFrame, output_dir: Path) -> dict:
    X_new = normalize_array(df_test[INPUTS].values, INPUTS, norm_stats)
    y_true_n = normalize_array(df_test[OUTPUTS].values, OUTPUTS, norm_stats)
    y_true = df_test[OUTPUTS].values

    modele.eval()
    with torch.no_grad():
        y_pred_n = modele(torch.tensor(X_new)).numpy()

    mu_out = norm_stats.loc[OUTPUTS, "mean"].values
    sd_out = norm_stats.loc[OUTPUTS, "std"].values
    y_pred = y_pred_n * sd_out + mu_out

    fig, axes = plt.subplots(1, len(OUTPUTS), figsize=(6*len(OUTPUTS), 6), squeeze=False)
    axes = axes.flatten()
    metrics = {}
    for i, (ax, col) in enumerate(zip(axes, OUTPUTS)):
        y_r, y_p = y_true[:, i], y_pred[:, i]
        ax.scatter(y_r, y_p, alpha=0.4, s=8)
        lim = max(abs(y_r).max(), abs(y_p).max())
        ax.plot([-lim, lim], [-lim, lim], "r--", lw=1, label="prédiction parfaite")
        ax.set_xlabel(f"{col} réel (physique)")
        ax.set_ylabel(f"{col} prédit (physique)")

        mse_norm = ((y_pred_n[:, i] - y_true_n[:, i]) ** 2).mean()
        r2 = 1 - mse_norm / y_true_n[:, i].var()

        ax.set_title(f"{col}\nMSE (norm)={mse_norm:.2e}  |  R²={r2:.3f}")
        ax.legend(); ax.grid(True)
        metrics[col] = {"mse_norm": float(mse_norm), "r2": float(r2)}

    fig.suptitle("Test sur toutes les données test du dataset", fontsize=14)
    plt.tight_layout()
    plt.savefig(output_dir / "test_predictions.png", dpi=150, bbox_inches="tight")
    plt.close()
    return metrics


@dataclass
class RolloutResult:
    U: np.ndarray
    U_reel: np.ndarray
    A: float
    omega: float


def _biais_repos(modele, mu_in, sd_in, mu_out, sd_out, cfg: Config):
    # Sortie réseau pour une entrée nulle, soustraite au rollout pour que la
    # zone au repos reste à 0.
    Xz = (np.zeros((len(cfg.nodes), len(mu_in)), dtype=np.float32) - mu_in) / sd_in
    with torch.no_grad():
        return (modele(torch.tensor(Xz)).numpy() * sd_out + mu_out)[0]


def _autoregressive_rollout(modele, U_reel, input_fields, mu_in, sd_in, mu_out, sd_out,
                             biais_repos, A, omega, cfg: Config) -> np.ndarray:
    # Réutilisée par run_rollout et le benchmark. Les M_BACK*ndt premiers pas
    # viennent de U_reel (historique nécessaire avant la 1re prédiction),
    # au-delà le champ est généré de façon autorégressive par le réseau.
    history_needed = cfg.M_BACK * cfg.ndt
    U = np.zeros((cfg.Nt + 1, cfg.Ntot))
    for m in range(history_needed + 1):
        U[m] = U_reel[m]

    for n in range(history_needed, cfg.Nt - cfg.N_FWD*cfg.ndt + 1, cfg.N_FWD*cfg.ndt):
        m_list = [n - lag*cfg.ndt for lag in range(cfg.M_BACK)]
        X = (build_window(m_list, lambda m: U[m], input_fields, cfg) - mu_in) / sd_in
        with torch.no_grad():
            sortie = modele(torch.tensor(X)).numpy()
        deltas = sortie * sd_out + mu_out - biais_repos

        for h in range(1, cfg.N_FWD + 1):
            s = n + h*cfg.ndt
            U[s, cfg.nodes] = U[n, cfg.nodes] + deltas[:, h-1]
            U[s, :cfg.i_left+1] = 0.0
            U[s, cfg.i_right:] = u_right_val(A, omega, s*cfg.dt)

            if cfg.SMOOTH_ALPHA > 0:
                # Lissage laplacien : atténue le bruit haute fréquence sans amortir l'onde.
                j0, j1 = cfg.i_left + 1, cfg.i_right
                lap = U[s, j0-1:j1-1] - 2*U[s, j0:j1] + U[s, j0+1:j1+1]
                U[s, j0:j1] += cfg.SMOOTH_ALPHA * lap

    return U


def run_rollout(modele, FIELDS, input_fields, norm_stats, INPUTS, OUTPUTS, cfg: Config) -> RolloutResult:
    A = cfg.AMPLITUDES[cfg.ROLLOUT_A_IDX]
    omega = cfg.PULSATIONS[cfg.ROLLOUT_OMEGA_IDX]
    U_reel = FIELDS[(A, omega)]

    mu_in = norm_stats.loc[INPUTS, "mean"].values.astype(np.float32)
    sd_in = norm_stats.loc[INPUTS, "std"].values.astype(np.float32)
    mu_out = norm_stats.loc[OUTPUTS, "mean"].values.astype(np.float32)
    sd_out = norm_stats.loc[OUTPUTS, "std"].values.astype(np.float32)

    biais_repos = _biais_repos(modele, mu_in, sd_in, mu_out, sd_out, cfg)
    U = _autoregressive_rollout(modele, U_reel, input_fields, mu_in, sd_in, mu_out, sd_out,
                                 biais_repos, A, omega, cfg)
    return RolloutResult(U=U, U_reel=U_reel, A=A, omega=omega)


def l2_rel(pred, true, eps=1e-12):
    return np.linalg.norm(pred - true) / (np.linalg.norm(true) + eps)


def smape(pred, true):
    m = true != 0
    return np.mean(2*np.abs(true[m] - pred[m]) / (np.abs(true[m]) + np.abs(pred[m])))


def compute_errors(rollout: RolloutResult, cfg: Config):
    steps = np.arange(2*cfg.ndt, cfg.Nt + 1, cfg.ndt)
    t_axis = steps * cfg.dt
    nodes = cfg.nodes
    U, U_reel = rollout.U, rollout.U_reel
    l2_list = [l2_rel(U[k, nodes], U_reel[k, nodes]) for k in steps]
    linf_list = [np.max(np.abs(U[k, nodes] - U_reel[k, nodes])) for k in steps]
    smape_list = [100.0 * smape(U[k, nodes], U_reel[k, nodes]) for k in steps]
    return t_axis, l2_list, linf_list, smape_list


def plot_rollout_error(t_axis, l2_list, linf_list, output_dir: Path):
    plt.figure(figsize=(9, 5))
    plt.plot(t_axis, l2_list, "o-", ms=3, label="erreur L2 relative")
    plt.plot(t_axis, linf_list, "s-", ms=3, label="erreur max absolue (Linf)")
    plt.yscale("log")
    plt.xlabel("t"); plt.ylabel("erreur"); plt.grid(True, which="both"); plt.legend()
    plt.title("Erreur du rollout en fonction du temps")
    plt.savefig(output_dir / "erreur_temps.png", dpi=150, bbox_inches="tight")
    plt.close()


def plot_smape(t_axis, smape_list, output_dir: Path):
    plt.figure(figsize=(9, 5))
    plt.plot(t_axis, smape_list, "s-", ms=3, label="sMAPE")
    plt.xlabel("t"); plt.ylabel("erreur (%)"); plt.grid(True); plt.legend()
    plt.title("sMAPE du rollout en fonction du temps")
    plt.savefig(output_dir / "smape_temps.png", dpi=150, bbox_inches="tight")
    plt.close()


def make_rollout_animation(rollout: RolloutResult, cfg: Config, output_dir: Path):
    U, U_reel = rollout.U, rollout.U_reel
    nodes = cfg.nodes
    x = np.linspace(0, cfg.L, cfg.Nx)
    frames = np.arange(0, cfg.Nt + 1, cfg.ndt)

    fig_anim, (axA, axB) = plt.subplots(2, 1, figsize=(9, 7), sharex=True)

    ligne_reel, = axA.plot([], [], "r", lw=2, label="réel")
    ligne_pred, = axA.plot([], [], "b--", lw=2, label="prédit")
    ymax = np.abs(U_reel[:, nodes]).max() * 1.2
    axA.set_xlim(0, cfg.L); axA.set_ylim(-ymax, ymax)
    axA.set_ylabel("u"); axA.legend(loc="upper right"); axA.grid(True)

    ligne_err, = axB.plot([], [], "k", lw=1.5, label="|prédit - réel|")
    err_max = max(np.max([np.abs(U[m, nodes] - U_reel[m, nodes]).max() for m in frames]) * 1.2, 1e-9)
    axB.set_xlim(0, cfg.L); axB.set_ylim(0, err_max)
    axB.set_xlabel("x"); axB.set_ylabel("erreur absolue"); axB.legend(loc="upper right"); axB.grid(True)

    titre = fig_anim.suptitle("")

    def maj(m):
        ligne_reel.set_data(x, U_reel[m, nodes])
        ligne_pred.set_data(x, U[m, nodes])
        ligne_err.set_data(x, np.abs(U[m, nodes] - U_reel[m, nodes]))
        titre.set_text(f"Propagation de l'onde — t = {m*cfg.dt:.3f}  (pas {m})")
        return ligne_reel, ligne_pred, ligne_err, titre

    anim = animation.FuncAnimation(fig_anim, maj, frames=frames, interval=50, blit=False)
    anim.save(output_dir / "propagation_onde.gif", writer="pillow", fps=20, dpi=110)
    plt.close(fig_anim)


def plot_utt_uxx(rollout: RolloutResult, cfg: Config, output_dir: Path):
    # Vérification EDP : u_tt en fonction de u_xx, réel puis prédit.
    U, U_reel = rollout.U, rollout.U_reel
    i_left, i_right = cfg.i_left, cfg.i_right
    dt, dx, ndt, Nt, Ntot = cfg.dt, cfg.dx, cfg.ndt, cfg.Nt, cfg.Ntot

    ureel_tt = np.zeros((Nt, Ntot)); ureel_xx = np.zeros((Nt, Ntot))
    for n in range(1, Nt):
        u_prev, u_curr, u_next = U_reel[n-1], U_reel[n], U_reel[n+1]
        ureel_tt[n, i_left:i_right+1] = (u_next[i_left:i_right+1] - 2*u_curr[i_left:i_right+1] + u_prev[i_left:i_right+1]) / dt**2
        ureel_xx[n, i_left:i_right+1] = (u_curr[i_left-1:i_right] - 2*u_curr[i_left:i_right+1] + u_curr[i_left+1:i_right+2]) / dx**2

    snaps_reel = sorted({int(np.clip(round(frac * Nt), 1, Nt - 1)) for frac in (0.1, 0.2, 0.4)})
    plt.figure()
    for n in snaps_reel:
        plt.scatter(ureel_xx[n, i_left+1:i_right], ureel_tt[n, i_left+1:i_right], s=10, label=f"n = {n}")
    plt.xlabel("u_xx (réel)"); plt.ylabel("u_tt (réel)")
    plt.legend(); plt.grid(); plt.xlim(-10, 10); plt.ylim(-10, 10)
    plt.title("u_tt en fonction de u_xx (real)")
    plt.savefig(output_dir / "utt_uxx_reel.png", dpi=150, bbox_inches="tight")
    plt.close()

    upred_tt = np.zeros((Nt, Ntot)); upred_xx = np.zeros((Nt, Ntot))
    for n in range(ndt, Nt - ndt + 1, ndt):
        u_prev, u_curr, u_next = U[n-ndt], U[n], U[n+ndt]
        upred_tt[n, i_left:i_right+1] = (u_next[i_left:i_right+1] - 2*u_curr[i_left:i_right+1] + u_prev[i_left:i_right+1]) / (ndt*dt)**2
        upred_xx[n, i_left:i_right+1] = (u_curr[i_left-1:i_right] - 2*u_curr[i_left:i_right+1] + u_curr[i_left+1:i_right+2]) / dx**2

    n_max_pred = ((Nt - ndt) // ndt) * ndt
    snaps_pred = sorted({int(np.clip(round(frac * Nt / ndt) * ndt, ndt, n_max_pred)) for frac in (0.01, 0.2, 0.3)})
    plt.figure()
    for n in snaps_pred:
        plt.scatter(upred_xx[n, i_left+1:i_right], upred_tt[n, i_left+1:i_right], s=10, label=f"n = {n}")
    plt.xlabel("u_xx (predit)"); plt.ylabel("u_tt (predit)")
    plt.grid(); plt.xlim(-10, 10); plt.ylim(-10, 10); plt.legend()
    plt.title("u_tt en fonction de u_xx (prediction)")
    plt.savefig(output_dir / "utt_uxx_predit.png", dpi=150, bbox_inches="tight")
    plt.close()


def chrono(fonction, n_repeat=15, n_warmup=3):
    for _ in range(n_warmup):
        fonction()
    durations = []
    for _ in range(n_repeat):
        t0 = time.perf_counter()
        fonction()
        durations.append(time.perf_counter() - t0)
    d = np.array(durations)
    return d.mean(), d.std(), float(np.median(d))


@dataclass
class BenchmarkResult:
    fd_time_med: float
    fd_time_std: float
    nn_time_med: float
    nn_time_std: float
    flops_per_call: float
    n_calls: int


def benchmark_inference(modele, FIELDS, input_fields, norm_stats, INPUTS, OUTPUTS,
                         rollout: RolloutResult, cfg: Config) -> BenchmarkResult:
    A, omega, U_reel = rollout.A, rollout.omega, rollout.U_reel

    mu_in = norm_stats.loc[INPUTS, "mean"].values.astype(np.float32)
    sd_in = norm_stats.loc[INPUTS, "std"].values.astype(np.float32)
    mu_out = norm_stats.loc[OUTPUTS, "mean"].values.astype(np.float32)
    sd_out = norm_stats.loc[OUTPUTS, "std"].values.astype(np.float32)
    biais_repos = _biais_repos(modele, mu_in, sd_in, mu_out, sd_out, cfg)

    def fd_once():
        return run_fd_simulation(A, omega, cfg)

    def rollout_once():
        return _autoregressive_rollout(modele, U_reel, input_fields, mu_in, sd_in, mu_out, sd_out,
                                        biais_repos, A, omega, cfg)

    fd_mean, fd_std, fd_med = chrono(fd_once)
    nn_mean, nn_std, nn_med = chrono(rollout_once)

    n_calls = len(range(cfg.M_BACK*cfg.ndt, cfg.Nt - cfg.N_FWD*cfg.ndt + 1, cfg.N_FWD*cfg.ndt))
    n_features = cfg.M_BACK * (2*cfg.SS + 1) * len(input_fields)

    from torch.utils.flop_counter import FlopCounterMode
    with FlopCounterMode(display=False) as fc:
        modele(torch.zeros((len(cfg.nodes), n_features)))
    flops_per_call = fc.get_total_flops() * n_calls

    print(f"FD (réel)   : {fd_med*1e3:7.3f} ms")
    print(f"NN (rollout): {nn_med*1e3:7.3f} ms  (±{nn_std*1e3:.3f})")
    print(f"speedup FD/NN = {fd_med/nn_med:.2f}x   (>1 = le réseau est plus rapide)")

    return BenchmarkResult(
        fd_time_med=fd_med, fd_time_std=float(fd_std),
        nn_time_med=nn_med, nn_time_std=float(nn_std),
        flops_per_call=flops_per_call, n_calls=n_calls,
    )


def export_resume(output_dir: Path, cfg: Config, method_name: str, df: pd.DataFrame, INPUTS, OUTPUTS,
                   train_result: TrainResult, tf_metrics: dict, rollout: RolloutResult,
                   bench: BenchmarkResult, errors):
    t_axis, l2_list, linf_list, smape_list = errors
    l2_final, l2_max = l2_list[-1], max(l2_list)
    linf_final, linf_max = linf_list[-1], max(linf_list)
    smape_final, smape_max = smape_list[-1], max(smape_list)

    with open(output_dir / "resume.txt", "w") as f:
        f.write("=====  RESUME DU RUN  =====\n\n")
        f.write(f"Methode         : {method_name}\n\n")

        f.write("--- Configuration ---\n")
        f.write(f"Grille          : Nt={cfg.Nt}, Nx={cfg.Nx}, SS={cfg.SS}, ndt={cfg.ndt}\n")
        f.write(f"M / N           : M_BACK={cfg.M_BACK}, N_FWD={cfg.N_FWD}\n")
        f.write(f"Rollout (A,w)   : A={rollout.A}, omega={rollout.omega}\n")
        f.write(f"Dataset         : {len(df):,} lignes\n")
        f.write(f"Features        : {len(INPUTS)} entrees, {len(OUTPUTS)} sortie(s)\n")
        for s in ["train", "val", "test"]:
            n = (df["split"] == s).sum()
            f.write(f"  split {s:5s}   : {n:>8,} lignes ({100*n/len(df):.1f} %)\n")
        f.write(f"Parametres NN   : {train_result.n_params:,}\n\n")

        f.write("--- Entrainement ---\n")
        f.write(f"Val minimale    : {train_result.meilleure_val:.6e}\n")
        for col, m in tf_metrics.items():
            f.write(f"{col:15s} : MSE (norm) = {m['mse_norm']:.4e} | R2 = {m['r2']:.4f}\n")
        f.write("\n")

        f.write("--- Temps d'execution ---\n")
        f.write(f"Entrainement ({cfg.N_EPOCHS} epochs)     : {train_result.train_time_s:.3f} s\n")
        f.write(f"Simulation reelle (FD, mediane)      : {bench.fd_time_med*1e3:.3f} ms\n")
        f.write(f"Rollout predit (NN, mediane)         : {bench.nn_time_med*1e3:.3f} ms  (+/-{bench.nn_time_std*1e3:.3f})\n")
        f.write(f"Speedup FD/NN                        : {bench.fd_time_med/bench.nn_time_med:.2f}\n")
        f.write(f"FLOPs reseau (rollout complet)        : {bench.flops_per_call:,.0f}\n\n")

        f.write("--- Erreurs du rollout ---\n")
        f.write(f"L2 relative  : finale = {l2_final:.4e}  |  max = {l2_max:.4e}\n")
        f.write(f"Linf absolue : finale = {linf_final:.4e}  |  max = {linf_max:.4e}\n")
        f.write(f"sMAPE (%)    : finale = {smape_final:.3f}  |  max = {smape_max:.3f}\n")

    print(f"Resume sauvegarde : {output_dir / 'resume.txt'}")


# Écritures concurrentes possibles (méthodes en parallèle) : flock sérialise
# le load/modifie/save de comparative_table.xlsx entre les process.
@contextmanager
def _xlsx_lock(xlsx_path: Path):
    lock_path = xlsx_path.with_suffix(xlsx_path.suffix + ".lock")
    with open(lock_path, "w") as f:
        fcntl.flock(f, fcntl.LOCK_EX)
        try:
            yield
        finally:
            fcntl.flock(f, fcntl.LOCK_UN)


def export_errors_to_xlsx(xlsx_path: Path, method_name: str, t_axis, l2_list):
    import openpyxl
    with _xlsx_lock(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

        header = [cell.value for cell in ws[1]]
        if method_name in header:
            col_err = header.index(method_name) + 1
            col_time = col_err - 1
        else:
            col_time = ws.max_column + 1
            col_err = col_time + 1
            ws.cell(row=1, column=col_time, value="t")
            ws.cell(row=1, column=col_err, value=method_name)

        # Efface l'existant avant d'écrire (évite les lignes fantômes d'un
        # run précédent qui aurait produit plus de points).
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_time, value=None)
            ws.cell(row=row, column=col_err, value=None)

        for i, (t, e) in enumerate(zip(t_axis, l2_list), start=2):
            ws.cell(row=i, column=col_time, value=float(t))
            ws.cell(row=i, column=col_err, value=float(e))

        wb.save(xlsx_path)
    print(f"{len(l2_list)} valeurs écrites dans {xlsx_path} (colonnes t / {method_name}).")


def export_timings_to_xlsx(xlsx_path: Path, method_name: str, train_time_s: float,
                            rollout_time_med_s: float, rollout_time_std_s: float,
                            fd_time_med_s: float, n_params: int, flops_per_call: float):
    import openpyxl
    with _xlsx_lock(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
        if "Timings" not in wb.sheetnames:
            ws = wb.create_sheet("Timings")
            ws.append(["method", "train_time_s", "rollout_time_median_s", "rollout_time_std_s",
                       "fd_time_median_s", "speedup_fd_over_nn", "n_params", "flops_per_rollout_call", "run_timestamp"])
        else:
            ws = wb["Timings"]

        speedup = fd_time_med_s / rollout_time_med_s if rollout_time_med_s else float("nan")
        row_values = [method_name, train_time_s, rollout_time_med_s, rollout_time_std_s,
                      fd_time_med_s, speedup, n_params, flops_per_call, datetime.now().isoformat(timespec="seconds")]

        target_row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == method_name:
                target_row = r
                break
        if target_row is None:
            target_row = ws.max_row + 1

        for col, val in enumerate(row_values, start=1):
            ws.cell(row=target_row, column=col, value=val)

        wb.save(xlsx_path)
    print(f"Temps écrits dans {xlsx_path} (feuille Timings, méthode {method_name}).")
