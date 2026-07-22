# ============================================================
# COMPARAISON DES 6 VARIANTES DE L'ABLATION "SUPPRESSION HAUTES FREQUENCES"
# (baseline, v1_smoothing, v2_noise_injection, v3_spectral_penalty,
#  v4_spectral_filter, v5_pushforward)
#
# Lit outputs_<variant>/metrics.json de chaque copie du script d'ablation
# (dossier parent de celui-ci) et produit :
#   - Comparatifs/comparative_table.xlsx : feuille "Resume" (une ligne par
#     metrique, une colonne par variante) + feuille "RolloutCurves" (paires
#     temps/RMSE par variante, meme convention que le fichier
#     Code_comparaison_des_inputs/Comparaisons/comparative_table.xlsx utilise
#     ailleurs dans le projet, mais dans un fichier LOCAL a cette ablation :
#     on n'ecrit pas dans le fichier de l'autre etude, cf. RESUME_CHANGEMENTS.md)
#   - Comparatifs/outputs/*.png : graphes de comparaison
#
# A executer APRES que les scripts d'ablation (baseline + v1..v5, dans le
# dossier parent) ont tourne au moins une fois. Chacun ecrit son propre
# outputs_<variant>/metrics.json ; les variantes pas encore lancees sont
# ignorees ici (avec un avertissement) plutot que de faire planter l'agregation
# -- tu peux donc lancer ce script apres seulement 1 ou 2 variantes pour un
# premier apercu, puis le relancer une fois toutes les variantes terminees.
# ============================================================
from pathlib import Path
import json

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR   = SCRIPT_DIR.parent
OUTPUT_DIR = SCRIPT_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

VARIANTS = ["baseline", "v1_smoothing", "v2_noise_injection",
            "v3_spectral_penalty", "v4_spectral_filter", "v5_pushforward"]

LABELS = {
    "baseline":             "Baseline (actuelle)",
    "v1_smoothing":         "v1 : smoothing renforce",
    "v2_noise_injection":   "v2 : noise injection HF",
    "v3_spectral_penalty":  "v3 : penalite spectrale",
    "v4_spectral_filter":   "v4 : filtre spectral",
    "v5_pushforward":       "v5 : pushforward renforce",
}

# Palette categorielle colorblind-safe, ordre fixe (jamais recycle) --
# memes 6 premieres teintes que la palette de reference du projet
# (deja utilisee dans Code_comparaison_des_inputs/Comparaisons/compare_methodes.py).
COLORS = {
    "baseline":             "#2a78d6",  # bleu
    "v1_smoothing":         "#008300",  # vert
    "v2_noise_injection":   "#e87ba4",  # magenta
    "v3_spectral_penalty":  "#eda100",  # jaune
    "v4_spectral_filter":   "#1baf7a",  # aqua
    "v5_pushforward":       "#eb6834",  # orange
}


def load_metrics(variant):
    path = ROOT_DIR / f"outputs_{variant}" / "metrics.json"
    if not path.exists():
        return None
    with open(path) as f:
        return json.load(f)


metrics   = {v: load_metrics(v) for v in VARIANTS}
dispo     = [v for v in VARIANTS if metrics[v] is not None]
manquants = [v for v in VARIANTS if metrics[v] is None]

print(f"Variantes disponibles : {dispo}")
if manquants:
    print(f"Variantes manquantes (ignorees) : {manquants}")
    print("  -> lance leur script (dans le dossier parent) puis relance cette agregation pour les inclure.")

if not dispo:
    raise SystemExit("Aucun metrics.json trouve. Lance au moins une variante avant l'agregation.")


# ============================================================
# TABLEAU COMPARATIF : comparative_table.xlsx
# ============================================================
xlsx_path = SCRIPT_DIR / "comparative_table.xlsx"
wb = openpyxl.Workbook()

# --- Feuille "Resume" : une colonne par methode, une ligne par metrique ---
ws = wb.active
ws.title = "Resume"

ROWS = [
    ("RMSE final",                  lambda m: m["rmse_final"]),
    ("RMSE max",                    lambda m: m["rmse_max"]),
    ("L2 relative final",           lambda m: m["l2_final"]),
    ("L2 relative max",             lambda m: m["l2_max"]),
    ("Linf final",                  lambda m: m["linf_final"]),
    ("Linf max",                    lambda m: m["linf_max"]),
    ("sMAPE final (%)",             lambda m: m["smape_final"]),
    ("sMAPE max (%)",               lambda m: m["smape_max"]),
    ("Seuil de divergence (RMSE)",  lambda m: m["divergence_rmse_threshold"]),
    ("Pas de divergence",           lambda m: m["divergence_step"] if m["divergence_step"] is not None else "aucune"),
    ("Temps de divergence (s)",     lambda m: m["divergence_time"] if m["divergence_time"] is not None else "aucune"),
    ("Pas de comparaison FFT",      lambda m: m["comparison_step"]),
    ("Energie HF au pas comp. (%)", lambda m: 100 * m["hf_energy_frac_at_comparison_step"]),
    ("Temps entrainement (s)",      lambda m: m["train_time_s"]),
    ("Nb parametres",               lambda m: m["n_params"]),
    ("Meilleure val loss",          lambda m: m["best_val_loss"]),
]

ws.cell(row=1, column=1, value="Metrique")
for j, v in enumerate(dispo, start=2):
    ws.cell(row=1, column=j, value=LABELS[v])

for i, (nom, fn) in enumerate(ROWS, start=2):
    ws.cell(row=i, column=1, value=nom)
    for j, v in enumerate(dispo, start=2):
        try:
            val = fn(metrics[v])
        except (KeyError, TypeError):
            val = None
        ws.cell(row=i, column=j, value=val)

# --- Feuille "RolloutCurves" : paires temps/RMSE par variante ---
ws2 = wb.create_sheet("RolloutCurves")
col = 1
for v in dispo:
    m = metrics[v]
    ws2.cell(row=1, column=col,     value="t")
    ws2.cell(row=1, column=col + 1, value=LABELS[v])
    for i, (t, e) in enumerate(zip(m["t_axis"], m["rmse"]), start=2):
        ws2.cell(row=i, column=col,     value=t)
        ws2.cell(row=i, column=col + 1, value=e)
    col += 2

wb.save(xlsx_path)
print(f"Tableau comparatif ecrit : {xlsx_path}")


# ============================================================
# GRAPHES DE COMPARAISON
# ============================================================
xs         = np.arange(len(dispo))
colors_bar = [COLORS[v] for v in dispo]

# --- RMSE(t), echelle log et lineaire ---
for echelle, suffixe in [("log", "log"), ("linear", "lin")]:
    plt.figure(figsize=(9, 5))
    for v in dispo:
        m = metrics[v]
        plt.plot(m["t_axis"], m["rmse"], "-", lw=1.8, color=COLORS[v], label=LABELS[v])
    plt.yscale(echelle)
    plt.xlabel("t"); plt.ylabel("RMSE")
    plt.title(f"RMSE du rollout — comparaison des methodes ({echelle})")
    plt.grid(True, which="both", alpha=0.3); plt.legend()
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / f"rmse_comparaison_{suffixe}.png", dpi=150, bbox_inches="tight")
    plt.close()

# --- Pas de divergence (bar chart ; plus haut = tient plus longtemps) ---
plt.figure(figsize=(8, 5))
steps_div = [metrics[v]["divergence_step"] for v in dispo]
Nt_max    = max(metrics[v]["steps"][-1] for v in dispo)
heights   = [s if s is not None else Nt_max for s in steps_div]
plt.bar(xs, heights, color=colors_bar)
for x, s, h in zip(xs, steps_div, heights):
    label = str(s) if s is not None else "pas de\ndivergence"
    plt.text(x, h + Nt_max * 0.01, label, ha="center", va="bottom", fontsize=8)
plt.xticks(xs, [LABELS[v] for v in dispo], rotation=20, ha="right")
plt.ylabel(f"pas de rollout (n) ou RMSE > {metrics[dispo[0]]['divergence_threshold_frac']} x A")
plt.title("Pas de divergence par methode (plus haut = mieux)")
plt.grid(axis="y", alpha=0.3)
plt.tight_layout()
plt.savefig(OUTPUT_DIR / "divergence_step_comparaison.png", dpi=150, bbox_inches="tight")
plt.close()

# --- Spectre de l'erreur au pas de comparaison ---
plt.figure(figsize=(9, 5))
for v in dispo:
    m = metrics[v]
    plt.plot(m["fft_freqs"], m["fft_power"], "-", lw=1.5, color=COLORS[v], label=LABELS[v])
plt.yscale("log")
plt.xlabel("frequence spatiale (1/m)"); plt.ylabel("puissance |FFT(erreur)|^2")
plt.title(f"Spectre de l'erreur au pas de comparaison (n={metrics[dispo[0]]['comparison_step']})")
plt.grid(True, which="both", alpha=0.3); plt.legend()
plt.tight_layout()
plt.savefig(OUTPUT_DIR / "spectre_comparaison.png", dpi=150, bbox_inches="tight")
plt.close()

# --- Energie haute frequence au pas de comparaison (bar chart) ---
plt.figure(figsize=(8, 5))
hf = [100 * metrics[v]["hf_energy_frac_at_comparison_step"] for v in dispo]
plt.bar(xs, hf, color=colors_bar)
plt.xticks(xs, [LABELS[v] for v in dispo], rotation=20, ha="right")
plt.ylabel("energie haute frequence (%) au pas de comparaison")
plt.title("Energie haute frequence (> 0.5 x Nyquist) par methode (plus bas = mieux)")
plt.grid(axis="y", alpha=0.3)
plt.tight_layout()
plt.savefig(OUTPUT_DIR / "hf_energy_comparaison.png", dpi=150, bbox_inches="tight")
plt.close()

# --- RMSE final vs max (bar chart groupe) ---
plt.figure(figsize=(8, 5))
width  = 0.35
rmse_f = [metrics[v]["rmse_final"] for v in dispo]
rmse_m = [metrics[v]["rmse_max"] for v in dispo]
plt.bar(xs - width / 2, rmse_f, width, label="RMSE final", color=colors_bar)
plt.bar(xs + width / 2, rmse_m, width, label="RMSE max", color=colors_bar, alpha=0.5)
plt.xticks(xs, [LABELS[v] for v in dispo], rotation=20, ha="right")
plt.ylabel("RMSE")
plt.title("RMSE final vs max par methode")
plt.legend(); plt.grid(axis="y", alpha=0.3)
plt.tight_layout()
plt.savefig(OUTPUT_DIR / "rmse_final_max_comparaison.png", dpi=150, bbox_inches="tight")
plt.close()

print(f"Graphes de comparaison sauvegardes dans {OUTPUT_DIR}")
