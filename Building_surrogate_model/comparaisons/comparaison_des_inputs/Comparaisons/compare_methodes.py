# Lit comparative_table.xlsx (erreur vs temps par méthode + feuille
# "Timings") et produit les graphes de comparaison dans Comparaisons/outputs/.
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

xlsx_path = SCRIPT_DIR / "comparative_table.xlsx"

# Groupes de comparaison suspendus (méthodes correspondantes non lancées dans
# commun.job) -- pas supprimés, décommenter en même temps que les méthodes
# dans commun.job pour les réactiver.
# METHODES = ["U", "U_Ut", "Ut_Uxx", "U_Ut_Uxx"]
# PF_EXPERIMENT = ["U", "U_pf40"]
# PF_EXPERIMENT_UT_UXX = ["Ut_Uxx", "Ut_Uxx_pf40"]

# Groupe actif : seul U_Ut_Uxx, U_pf40 et U_Ut_Uxx_pf40 tournent actuellement.
PF_EXPERIMENT_U_UT_UXX = ["U_Ut_Uxx", "U_pf40", "U_Ut_Uxx_pf40"]
# Couleur fixe par méthode sur tous les graphes du dossier.
METHOD_COLORS = {
    "U":               "#2a78d6",
    "U_Ut":            "#008300",
    "Ut_Uxx":          "#eb6834",
    "U_Ut_Uxx":        "#4a3aa7",
    "U_pf40":          "#c0392b",
    "Ut_Uxx_pf40":     "#8a5a00",
    "U_Ut_Uxx_pf40":   "#8a2a8a",
}

wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb.active

headers = [c.value for c in ws[1]]
print("Colonnes trouvées :", headers)


def col_index(nom):
    return headers.index(nom)   # 0-based


def extraire(nom):
    ci = col_index(nom)
    ci_time = ci - 1                     # la colonne temps est juste à gauche
    ts, es = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[ci_time]
        e = row[ci]
        if t is None or e is None:
            continue
        ts.append(float(t))
        es.append(float(e))
    return ts, es


def tracer(colonnes, nom_fichier, titre):
    colonnes_dispo = [c for c in colonnes if c in headers]
    manquantes = [c for c in colonnes if c not in headers]
    if manquantes:
        print(f"Colonnes absentes du xlsx (ignorées) : {manquantes}")
    if not colonnes_dispo:
        print(f"Aucune colonne disponible pour {nom_fichier}, graphe non généré.")
        return

    for echelle, suffixe, label in [("log", "log", "échelle log"), ("linear", "lin", "échelle linéaire")]:
        plt.figure(figsize=(9, 5))
        for nom in colonnes_dispo:
            ts, es = extraire(nom)
            plt.plot(ts, es, "o-", ms=3, label=nom, color=METHOD_COLORS.get(nom))
        plt.yscale(echelle)
        plt.xlabel("t"); plt.ylabel("erreur")
        plt.title(f"{titre} ({label})")
        plt.grid(True, which="both"); plt.legend()
        plt.tight_layout()
        plt.savefig(OUTPUT_DIR / f"{nom_fichier}_{suffixe}.png", dpi=150, bbox_inches="tight")
        plt.close()


# tracer(METHODES, "comparaison_4modeles", "Erreur en fonction du temps (4 méthodes)")
# tracer(PF_EXPERIMENT, "comparaison_pf40", "Effet de la profondeur du pushforward (U vs U_pf40)")
# tracer(PF_EXPERIMENT_UT_UXX, "comparaison_pf40_ut_uxx",
#        "Effet de la profondeur du pushforward (Ut_Uxx vs Ut_Uxx_pf40)")
tracer(PF_EXPERIMENT_U_UT_UXX, "comparaison_pf40_u_ut_uxx",
       "Effet de la profondeur du pushforward (U_Ut_Uxx vs U_pf40 vs U_Ut_Uxx_pf40)")


def extraire_timings():
    if "Timings" not in wb.sheetnames:
        return {}
    ws_t = wb["Timings"]
    header_t = [c.value for c in ws_t[1]]
    data = {}
    for row in ws_t.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        data[row[0]] = dict(zip(header_t, row))
    return data


timings = extraire_timings()
methodes_dispo = [m for m in PF_EXPERIMENT_U_UT_UXX if m in timings]

if methodes_dispo:
    colors = [METHOD_COLORS.get(m) for m in methodes_dispo]

    plt.figure(figsize=(7, 5))
    valeurs = [timings[m]["train_time_s"] for m in methodes_dispo]
    plt.bar(methodes_dispo, valeurs, color=colors)
    plt.ylabel("Temps d'entraînement (s)")
    plt.title("Temps d'entraînement par méthode")
    plt.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "training_time_comparison.png", dpi=150, bbox_inches="tight")
    plt.close()

    plt.figure(figsize=(7, 5))
    valeurs = [timings[m]["rollout_time_median_s"] * 1e3 for m in methodes_dispo]
    erreurs = [timings[m]["rollout_time_std_s"] * 1e3 for m in methodes_dispo]
    plt.bar(methodes_dispo, valeurs, yerr=erreurs, capsize=4, color=colors)
    plt.ylabel("Temps de rollout (ms, médiane ± std)")
    plt.title("Temps de rollout / inférence par méthode")
    plt.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "rollout_time_comparison.png", dpi=150, bbox_inches="tight")
    plt.close()

    plt.figure(figsize=(7, 5))
    valeurs = [timings[m]["speedup_fd_over_nn"] for m in methodes_dispo]
    plt.bar(methodes_dispo, valeurs, color=colors)
    plt.axhline(1.0, color="black", lw=1, ls="--")
    plt.ylabel("Speedup FD / NN (> 1 : le réseau est plus rapide)")
    plt.title("Speedup rollout réseau vs simulation FD")
    plt.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "speedup_comparison.png", dpi=150, bbox_inches="tight")
    plt.close()

    print(f"Graphes de temps sauvegardés dans {OUTPUT_DIR}")
else:
    print("Feuille 'Timings' absente ou vide -- graphes de temps non générés "
          "(normal avant le premier run des 4 méthodes).")

print(f"Plots sauvegardés dans {OUTPUT_DIR}")
