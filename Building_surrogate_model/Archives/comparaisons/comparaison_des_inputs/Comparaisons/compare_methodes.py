# Reads comparative_table.xlsx (error vs time per method + "Timings"
# sheet) and produces the comparison plots in Comparaisons/outputs/.
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

xlsx_path = SCRIPT_DIR / "comparative_table.xlsx"

# Comparison groups on hold (corresponding methods not launched in
# commun.job) -- not deleted, uncomment together with the methods
# in commun.job to reactivate them.
# METHODES = ["U", "U_Ut", "Ut_Uxx", "U_Ut_Uxx"]
# PF_EXPERIMENT = ["U", "U_pf40"]
# PF_EXPERIMENT_UT_UXX = ["Ut_Uxx", "Ut_Uxx_pf40"]

# Active group: only U_Ut_Uxx, U_pf40 and U_Ut_Uxx_pf40 currently run.
PF_EXPERIMENT_U_UT_UXX = ["U_Ut_Uxx", "U_pf40", "U_Ut_Uxx_pf40"]
# Fixed color per method across all plots in this folder.
METHOD_COLORS = {
    "U":               "#2a78d6",
    "U_Ut":            "#008300",
    "Ut_Uxx":          "#eb6834",
    "U_Ut_Uxx":        "#4a3aa7",
    "U_pf40":          "#c0392b",
    "Ut_Uxx_pf40":     "#8a5a00",
    "U_Ut_Uxx_pf40":   "#8a2a8a",
}

wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb.active

headers = [c.value for c in ws[1]]
print("Columns found:", headers)


def col_index(nom):
    return headers.index(nom)   # 0-based


def extraire(nom):
    ci = col_index(nom)
    ci_time = ci - 1                     # the time column is right to the left
    ts, es = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[ci_time]
        e = row[ci]
        if t is None or e is None:
            continue
        ts.append(float(t))
        es.append(float(e))
    return ts, es


def tracer(colonnes, nom_fichier, titre):
    colonnes_dispo = [c for c in colonnes if c in headers]
    manquantes = [c for c in colonnes if c not in headers]
    if manquantes:
        print(f"Columns missing from the xlsx (ignored): {manquantes}")
    if not colonnes_dispo:
        print(f"No column available for {nom_fichier}, plot not generated.")
        return

    for echelle, suffixe, label in [("log", "log", "log scale"), ("linear", "lin", "linear scale")]:
        plt.figure(figsize=(9, 5))
        for nom in colonnes_dispo:
            ts, es = extraire(nom)
            plt.plot(ts, es, "o-", ms=3, label=nom, color=METHOD_COLORS.get(nom))
        plt.yscale(echelle)
        plt.xlabel("t"); plt.ylabel("error")
        plt.title(f"{titre} ({label})")
        plt.grid(True, which="both"); plt.legend()
        plt.tight_layout()
        plt.savefig(OUTPUT_DIR / f"{nom_fichier}_{suffixe}.png", dpi=150, bbox_inches="tight")
        plt.close()


# tracer(METHODES, "comparaison_4modeles", "Error over time (4 methods)")
# tracer(PF_EXPERIMENT, "comparaison_pf40", "Effect of pushforward depth (U vs U_pf40)")
# tracer(PF_EXPERIMENT_UT_UXX, "comparaison_pf40_ut_uxx",
#        "Effect of pushforward depth (Ut_Uxx vs Ut_Uxx_pf40)")
tracer(PF_EXPERIMENT_U_UT_UXX, "comparaison_pf40_u_ut_uxx",
       "Effect of pushforward depth (U_Ut_Uxx vs U_pf40 vs U_Ut_Uxx_pf40)")


def extraire_timings():
    if "Timings" not in wb.sheetnames:
        return {}
    ws_t = wb["Timings"]
    header_t = [c.value for c in ws_t[1]]
    data = {}
    for row in ws_t.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        data[row[0]] = dict(zip(header_t, row))
    return data


timings = extraire_timings()
methodes_dispo = [m for m in PF_EXPERIMENT_U_UT_UXX if m in timings]

if methodes_dispo:
    colors = [METHOD_COLORS.get(m) for m in methodes_dispo]

    plt.figure(figsize=(7, 5))
    valeurs = [timings[m]["train_time_s"] for m in methodes_dispo]
    plt.bar(methodes_dispo, valeurs, color=colors)
    plt.ylabel("Training time (s)")
    plt.title("Training time per method")
    plt.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "training_time_comparison.png", dpi=150, bbox_inches="tight")
    plt.close()

    plt.figure(figsize=(7, 5))
    valeurs = [timings[m]["rollout_time_median_s"] * 1e3 for m in methodes_dispo]
    erreurs = [timings[m]["rollout_time_std_s"] * 1e3 for m in methodes_dispo]
    plt.bar(methodes_dispo, valeurs, yerr=erreurs, capsize=4, color=colors)
    plt.ylabel("Rollout time (ms, median ± std)")
    plt.title("Rollout / inference time per method")
    plt.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "rollout_time_comparison.png", dpi=150, bbox_inches="tight")
    plt.close()

    plt.figure(figsize=(7, 5))
    valeurs = [timings[m]["speedup_fd_over_nn"] for m in methodes_dispo]
    plt.bar(methodes_dispo, valeurs, color=colors)
    plt.axhline(1.0, color="black", lw=1, ls="--")
    plt.ylabel("Speedup FD / NN (> 1: the network is faster)")
    plt.title("Network rollout speedup vs FD simulation")
    plt.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "speedup_comparison.png", dpi=150, bbox_inches="tight")
    plt.close()

    print(f"Timing plots saved to {OUTPUT_DIR}")
else:
    print("'Timings' sheet missing or empty -- timing plots not generated "
          "(normal before the first run of the 4 methods).")

print(f"Plots saved to {OUTPUT_DIR}")
