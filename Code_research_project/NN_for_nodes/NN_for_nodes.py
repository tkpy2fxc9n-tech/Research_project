# ============================================================
# IMPORTS (tout regroupé ici)
# ============================================================
from pathlib import Path
import time
from itertools import product

import numpy as np
from numpy.lib.stride_tricks import sliding_window_view
import pandas as pd

import matplotlib
matplotlib.use("Agg")            # backend non-interactif : aucune fenêtre ne s'ouvre
import matplotlib.pyplot as plt
import matplotlib.animation as animation

import torch
import torch.nn as nn
from torch.utils.data import DataLoader, TensorDataset

SEED = 0
torch.manual_seed(SEED)
np.random.seed(SEED)

# Dossier où sont sauvegardés tous les graphes
SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)


# ============================================================
# PARAMÈTRES
# ============================================================
# Physique
E   = 1
rho = 2
L   = 1

# Grille espace / temps
Nt   = 500
Nx   = 100
SS   = 11                       
Ntot = Nx + 2*SS                  

nodes = np.arange(SS, Ntot-SS)   

t_end = 5
dt = t_end / Nt
dx = L / (Nx - 1)
CFL = dt/dx * np.sqrt(E/rho)

# Position physique de chaque nœud (feature statique, indépendante de t) :
# donne au réseau un moyen de distinguer le voisinage d'un bord (encastré en
# x=0, piloté en x=L) du volume, plutôt que de supposer la physique
# invariante par translation partout.
x_nodes = np.linspace(0, L, Nx).astype(np.float32)

# Rollout : M instants passés -> N instants futurs (espacés de ndt pas).
# N_FWD determine aussi le pas d'avancement du rollout : le bloc suivant
# repart de l'horizon n+N_FWD*ndt, donc TOUJOURS de la prediction la MOINS
# fiable du bloc (l'horizon le plus loin a le R2 le plus faible). Un N_FWD
# elevé (essaye a 9 : R2 tombe de 0.98 a 0.94 entre le 1er et le 9e horizon)
# fait donc reboucler le rollout sur sa pire prediction a chaque bloc, ce qui
# le fait decrocher de la vraie onde au bout de quelques blocs (visible dans
# outputs/propagation_onde.gif). On garde N_FWD=M_BACK=3 : le rollout
# repart toujours de l'horizon le plus fiable (R2~0.99).
ndt    = 3
M_BACK = 3     # niveaux temporels en entrée  : t, t-ndt, ..., t-(M_BACK-1)*ndt
N_FWD  = 3     # horizons de sortie           : n+ndt, n+2ndt, ..., n+N_FWD*ndt

def jlabel(k):                    # libellé de colonne pour le voisin j+k
    return "j" if k == 0 else f"j{k:+d}"

# Jeu de simulations : N valeurs d'amplitude × N valeurs de pulsation
# N=3 (au lieu de 6) : dataset devenu trop volumineux en memoire (~9 Go pour
# le seul df, cf. discussion) -> on reduit le nombre de simulations plutot
# que la resolution physique (Nt, Nx), pour ne pas toucher au CFL.
N = 3
AMPLITUDES = np.linspace(0.005, 0.1, N).round(3).tolist()
PULSATIONS = np.linspace(3, 10, N).round(1).tolist()

# Sous-echantillonnage temporel du dataset (pas de la physique) : les pas de
# temps consecutifs sont tres correles, on ne garde qu'un pas sur TIME_STRIDE
# lors de la construction des lignes du dataset -> reduit le nombre de lignes
# sans rien changer a la simulation (u_storage) ni au rollout.
TIME_STRIDE = 3

# Garde-fou numérique (entraînement pushforward + rollout) : la physique
# (équation d'onde linéaire, bord droit borné par l'amplitude A) ne justifie
# jamais une amplitude de plusieurs ordres de grandeur au-dessus de la plus
# grande impulsion imposée. u_xx (÷dx² ≈ ×1e4) amplifie le moindre bruit sur
# le champ reconstruit -> sans garde-fou, une divergence peut s'auto-entretenir
# sur des dizaines de blocs successifs. On écrête plutôt que de laisser
# l'erreur exploser.
CLAMP_AMP = 10.0 * max(AMPLITUDES)


# ============================================================
# GÉNÉRATION DES DONNÉES
# Pour chaque (shape, A, omega) : on simule l'onde par différences finies,
# puis on en extrait les entrées (M_BACK niveaux) et sorties (N_FWD horizons).
# Plusieurs familles de forme de CL (shape) sont utilisées pour éviter que le
# réseau ne mémorise la seule forme "gauss" au lieu d'apprendre la physique
# locale (qui doit être valable quelle que soit l'excitation au bord).
# ============================================================
all_dfs = []
FIELDS = {}          # champ complet u(t,x) de chaque simulation, pour le pushforward

SHAPES = ["gauss", "sine_burst", "step_smooth"]

def u_right_val(shape, A, omega, t):
    sigma = np.interp(omega, [1.0, 10.0], [0.15, 0.07])
    t0    = 4.0 * sigma
    if shape == "gauss":
        # Impulsion gaussienne isolée, revient à zéro.
        return A * np.exp(-((t - t0) / sigma) ** 2)
    elif shape == "sine_burst":
        # Salve sinusoïdale de quelques périodes, enveloppe cosinus (0 aux bords).
        n_cycles = 3.0
        t1 = n_cycles * 2.0 * np.pi / omega
        if t <= 0.0 or t >= t1:
            return 0.0
        fenetre = 0.5 * (1.0 - np.cos(2.0 * np.pi * t / t1))
        return A * fenetre * np.sin(omega * t)
    elif shape == "step_smooth":
        # Montée lisse (cosinus) de 0 à A puis maintien -> charge soutenue,
        # régime qualitativement différent d'une impulsion qui repart à zéro.
        t1 = 4.0 * sigma
        if t <= 0.0:
            return 0.0
        elif t >= t1:
            return A
        else:
            return A * 0.5 * (1.0 - np.cos(np.pi * t / t1))
    else:
        raise ValueError(f"forme d'onde inconnue : {shape}")

for shape, A, omega in product(SHAPES, AMPLITUDES, PULSATIONS):

    u    = np.zeros(Ntot)
    u_1  = np.zeros(Ntot)
    u_xx = np.zeros(Ntot)

    u_storage    = np.zeros((Nt + 1, Ntot))
    u_xx_storage = np.zeros((Nt,     Ntot))
    u_storage[0] = u.copy()

    i_left  = SS
    i_right = Ntot - SS

    # --- Boucle de simulation sur la grille étendue ---
    for n in range(Nt):
        t = n * dt
        u_xx[i_left:i_right+1] = (u[i_left-1:i_right] - 2*u[i_left:i_right+1] + u[i_left+1:i_right+2]) / dx**2

        u_new = np.zeros(Ntot)
        u_new[i_left:i_right+1] = (2.0 * u[i_left:i_right+1] - u_1[i_left:i_right+1] + CFL**2 * (u[i_left-1:i_right] - 2.0 * u[i_left:i_right+1] + u[i_left+1:i_right+2]))
        u_new[:i_left+1] = 0.0
        u_new[i_right:] = u_right_val(shape, A, omega, t + dt)

        u_xx_storage[n] = u_xx
        u_1 = u.copy()
        u   = u_new
        u_storage[n + 1] = u.copy()

    FIELDS[(shape, A, omega)] = u_storage    # on garde le champ complet pour le pushforward

    # --- Construction du dataset : M_BACK instants passés, voisinage ±SS ---
    rows = []

    for n in range(M_BACK*ndt, Nt - N_FWD*ndt + 1, TIME_STRIDE):

        row = {"shape": shape, "A": A, "omega": omega, "n_step": n}

        # ---- ENTRÉES : M_BACK niveaux en arrière ----
        for lag in range(M_BACK):
            m   = n - lag*ndt
            lab = "t" if lag == 0 else f"t-{lag}ndt"
            u_lag    = u_storage[m]
            udot_lag = (u_storage[m] - u_storage[m - ndt]) / (ndt * dt)
            uxx_lag  = u_xx_storage[m]
            for k in range(-SS, SS + 1):
                # float32 : le dataset (millions de lignes) est le principal
                # poste memoire, contrairement a la simulation (u_storage).
                row[f"u({lab},{jlabel(k)})"]     = u_lag[nodes + k].astype(np.float32)
                row[f"u_dot({lab},{jlabel(k)})"] = udot_lag[nodes + k].astype(np.float32)
                row[f"u_xx({lab},{jlabel(k)})"]  = uxx_lag[nodes + k].astype(np.float32)

        # ---- Position physique (statique, un seul niveau) ----
        row["x_pos"] = x_nodes

        # ---- SORTIES : N_FWD horizons en avant ----
        for h in range(1, N_FWD + 1):
            row[f"delta_u@{h}ndt"] = (u_storage[n + h*ndt, nodes] - u_storage[n, nodes]).astype(np.float32)
        rows.append(pd.DataFrame(row))

    all_dfs.append(pd.concat(rows, ignore_index=True))

df = pd.concat(all_dfs, ignore_index=True)
print(df.head(0))
print(f"{len(df):,} lignes × {df.shape[1]} colonnes")


# ============================================================
# SPLIT train / val / test  +  NORMALISATION
# ============================================================
rng = np.random.default_rng(seed=42)

n_rows  = len(df)
n_train = int(0.90 * n_rows)
n_val   = int(0.05 * n_rows)
n_test  = n_rows - n_train - n_val

split_labels = np.array(["train"] * n_train + ["val"] * n_val + ["test"] * n_test)
rng.shuffle(split_labels)
df["split"] = split_labels

print("Distribution du split :")
for s in ["train", "val", "test"]:
    n = (df["split"] == s).sum()
    print(f"  {s:5s} : {n:>8,} lignes  ({100*n/len(df):.1f} %)")

# Colonnes d'entrée / sortie
OUTPUTS = [f"delta_u@{h}ndt" for h in range(1, N_FWD + 1)]
meta    = ["shape", "A", "omega", "n_step", "split"]
INPUTS  = [c for c in df.columns if c not in meta + OUTPUTS]

# Stats de normalisation calculées sur le TRAIN uniquement
train_mask = df["split"] == "train"
norm_stats = pd.DataFrame({
    "mean": df.loc[train_mask, INPUTS + OUTPUTS].mean(),
    "std" : df.loc[train_mask, INPUTS + OUTPUTS].std(),
})
norm_stats["std"] = norm_stats["std"].replace(0, 1)

# Application de la normalisation à tout le dataset (colonnes suffixées "_n").
# Un seul concat plutôt que ~140 insertions colonne par colonne (qui fragmentent
# le DataFrame et deviennent très lentes sur des millions de lignes).
# .astype(float32) : mean/std sont en float64 (précision du calcul des stats),
# sans ce cast la soustraction repasserait tout df_n en float64 et doublerait
# la mémoire déjà réduite par le passage en float32 à la construction.
df_n = ((df[INPUTS + OUTPUTS] - norm_stats["mean"]) / norm_stats["std"]).astype(np.float32)
df_n.columns = [c + "_n" for c in INPUTS + OUTPUTS]
df = pd.concat([df, df_n], axis=1)


# ============================================================
# HYPERPARAMÈTRES
# ============================================================
# Architecture du réseau
HIDDEN_SIZES = [64, 32, 16]

# Entraînement (le coût d'entraînement n'est pas contraint : seule la vitesse
# d'inférence du rollout face aux différences finies compte -> on peut se
# permettre beaucoup d'epochs et un pushforward plus large/fort).
LEARNING_RATE = 1e-3
N_EPOCHS      = 45
BATCH_SIZE    = 512

# Pushforward (stabilisation du rollout)
LAMBDA_PF   = 2.0   # poids de la loss pushforward (la loss données est déjà quasi
                     # parfaite en teacher forcing -> on privilégie la stabilité du rollout)
N_PF_GROUPS = 16    # nombre de simulations (champs complets) déroulées par batch
PF_WARMUP   = 5     # montée progressive de LAMBDA_PF sur les premières époques

# Loss "au repos" : force explicitement modele(entrée nulle) ~= 0 (delta nul).
# Sans ça, un biais résiduel se propage de façon uniforme dans les zones encore
# au repos à chaque bloc du rollout, ce qui contribue à la dérive/divergence
# à long horizon (corrigé après coup par `biais_repos`, mais mieux vaut que le
# réseau l'apprenne directement).
LAMBDA_REST = 0.1

# Robustification du rollout (distribution shift). Le rollout consomme ses propres
# prédictions ; u_xx (dérivée 2nde, ÷dx² ≈ ×1e4) amplifie le mode "damier" haute
# fréquence -> divergence. Deux parades :
#   1) NOISE_STD : on bruite les features (normalisées) à l'entraînement, donc le
#      réseau apprend à débruiter ses propres entrées.
#   2) SMOOTH_ALPHA : un pas de lissage Laplacien sur le champ prédit à chaque bloc
#      du rollout, qui amortit le damier (×(1-4α)) sans toucher l'onde (basse fréq).
NOISE_STD    = 0.10
SMOOTH_ALPHA = 0.20

print(f"Dataset : {len(df):,} lignes")
print(f"Splits  : {df['split'].value_counts().to_dict()}")


# ============================================================
# TENSEURS PyTorch + DataLoader d'entraînement
# ============================================================
INPUTS_N  = [f"{x}_n" for x in INPUTS]
OUTPUTS_N = [f"{x}_n" for x in OUTPUTS]

X_train = df.loc[df["split"] == "train", INPUTS_N].values.astype(np.float32)
y_train = df.loc[df["split"] == "train", OUTPUTS_N].values.astype(np.float32)

X_val   = df.loc[df["split"] == "val", INPUTS_N].values.astype(np.float32)
y_val   = df.loc[df["split"] == "val", OUTPUTS_N].values.astype(np.float32)

train_loader = DataLoader(TensorDataset(torch.tensor(X_train), torch.tensor(y_train)),
                          batch_size=BATCH_SIZE, shuffle=True)


# ============================================================
# OUTILS POUR LE PUSHFORWARD
#
# Le rollout diverge à cause du *distribution shift* : à l'entraînement le réseau
# voit des entrées propres (teacher forcing), mais au rollout il consomme ses
# propres prédictions, ré-injectées via u_xx (amplificateur de bruit). Le
# pushforward (Brandstetter et al. 2022) corrige ça : on déroule le réseau d'UN
# bloc sur sa propre prédiction (gradient détaché), puis on lui demande de revenir
# vers la vérité au bloc suivant.
# ============================================================
i_left, i_right = SS, Ntot - SS

# Moyennes / écarts-types, dans l'ordre EXACT de INPUTS / OUTPUTS
mu_in  = norm_stats.loc[INPUTS,  "mean"].values.astype(np.float32)
sd_in  = norm_stats.loc[INPUTS,  "std" ].values.astype(np.float32)
mu_out = norm_stats.loc[OUTPUTS, "mean"].values.astype(np.float32)
sd_out = norm_stats.loc[OUTPUTS, "std" ].values.astype(np.float32)

# Index de x_pos dans INPUTS : le "repos" (u=u_dot=u_xx=0) doit rester au vrai
# x de chaque nœud, pas être mis à x=0 (cf. remarques sur X_rest_n / Xz plus bas).
x_pos_idx = INPUTS.index("x_pos")



def u_xx_field(u):
    """u_xx sur la grille étendue (zéro hors [i_left, i_right]), comme le dataset."""
    out = np.zeros(Ntot)
    out[i_left:i_right+1] = (u[i_left-1:i_right] - 2*u[i_left:i_right+1] + u[i_left+1:i_right+2]) / dx**2
    return out

def build_window(m_list, field_at):
    # Vectorisé : pour off in [-SS, SS], u/udot/uxx[nodes+off] est exactement
    # la fenêtre glissante de largeur 2*SS+1 sur tout le tableau (nodes =
    # arange(SS, Ntot-SS)) -> sliding_window_view remplace la boucle Python
    # (M_BACK * (2*SS+1) itérations), appelée des dizaines de milliers de fois
    # pendant l'entraînement (pushforward) et le rollout.
    blocks = []
    for m in m_list:
        u    = field_at(m)
        udot = (u - field_at(m - ndt)) / (ndt * dt)
        uxx  = u_xx_field(u)
        win_u    = sliding_window_view(u,    2*SS + 1)   # (Nx, 2*SS+1)
        win_udot = sliding_window_view(udot, 2*SS + 1)
        win_uxx  = sliding_window_view(uxx,  2*SS + 1)
        blocks.append(np.stack([win_u, win_udot, win_uxx], axis=-1).reshape(len(nodes), -1))
    blocks.append(x_nodes.reshape(-1, 1))   # position physique (statique, un seul niveau)
    X = np.concatenate(blocks, axis=1).astype(np.float32)
    return (X - mu_in) / sd_in

def reconstruct(u_curr, n_curr, pred_norm, shape, A, omega):
    deltas = pred_norm * sd_out + mu_out
    champs = {}
    for h in range(1, N_FWD + 1):
        s = n_curr + h * ndt
        u = np.zeros(Ntot)
        u[nodes]     = np.clip(u_curr[nodes] + deltas[:, h-1], -CLAMP_AMP, CLAMP_AMP)
        u[:i_left+1] = 0.0
        u[i_right:]  = u_right_val(shape, A, omega, s * dt)
        champs[s] = u
    return champs


PF_SAMPLES = [(shape, A, omega, n)
              for (shape, A, omega) in FIELDS
              for n in range(M_BACK*ndt, Nt - 2*N_FWD*ndt + 1)]
print(f"Échantillons pushforward disponibles : {len(PF_SAMPLES):,}")


# ============================================================
# MODÈLE
# ============================================================
class Reseau(nn.Module):

    def __init__(self, n_inputs, n_outputs, hidden_sizes):
        super().__init__()
        couches = []
        taille_entree = n_inputs
        for taille in hidden_sizes:
            couches.append(nn.Linear(taille_entree, taille))
            couches.append(nn.GELU())
            taille_entree = taille
        couches.append(nn.Linear(taille_entree, n_outputs))
        self.reseau = nn.Sequential(*couches)

    def forward(self, x):
        return self.reseau(x)

modele = Reseau(n_inputs=len(INPUTS_N), n_outputs=len(OUTPUTS_N), hidden_sizes=HIDDEN_SIZES)
print(modele)
n_params = sum(p.numel() for p in modele.parameters())
print(f"\nNombre de paramètres : {n_params:,}")

criterion  = nn.MSELoss()
optimiseur = torch.optim.Adam(modele.parameters(), lr=LEARNING_RATE)
scheduler  = torch.optim.lr_scheduler.ReduceLROnPlateau(optimiseur, mode="min", factor=0.5, patience=10)

# Entrée/cible "au repos" pour LAMBDA_REST : entrée brute nulle (u_dot=u_xx=0
# partout) -> delta brut nul. En espace normalisé, entrée = -mu_in/sd_in et
# cible = -mu_out/sd_out (le zéro physique ne correspond pas au zéro normalisé).
#
# BUG corrigé : x_pos faisait partie des colonnes mises à zéro ci-dessus, donc
# cette régularisation ne forçait modele(repos)=0 qu'au nœud x=0. Ailleurs dans
# le domaine, rien n'empêchait un biais résiduel non nul — qui se réinjecte et
# s'accumule à chaque bloc du rollout (le décalage croissant visible dans
# propagation_onde.gif). On construit donc une ligne de repos PAR nœud, avec
# le vrai x_pos de ce nœud.
_rest_row = (-mu_in) / sd_in
X_rest_np = np.tile(_rest_row, (len(nodes), 1)).astype(np.float32)
X_rest_np[:, x_pos_idx] = (x_nodes - mu_in[x_pos_idx]) / sd_in[x_pos_idx]
X_rest_n = torch.tensor(X_rest_np)
y_rest_n = torch.tensor(((-mu_out) / sd_out).reshape(1, -1).astype(np.float32))

def pushforward_loss(n_groups):
    idxs = np.random.choice(len(PF_SAMPLES), n_groups, replace=False)
    groups = [PF_SAMPLES[i] for i in idxs]

    X1 = np.concatenate([build_window([n - lag*ndt for lag in range(M_BACK)], lambda m, U=FIELDS[(shape, A, omega)]: U[m]) for (shape, A, omega, n) in groups], axis=0,)
    with torch.no_grad():
        pred1 = modele(torch.tensor(X1)).numpy()

    nN = len(nodes)
    X2_list, tgt_list = [], []
    for j, (shape, A, omega, n) in enumerate(groups):
        U  = FIELDS[(shape, A, omega)]
        Up = reconstruct(U[n], n, pred1[j*nN:(j+1)*nN], shape, A, omega)
        field_at = lambda m, U=U, Up=Up: Up[m] if m in Up else U[m]

        nprime = n + N_FWD * ndt
        X2_list.append(build_window([nprime - lag*ndt for lag in range(M_BACK)], field_at))

        curr = Up[nprime][nodes]                                   
        tgt  = np.stack([U[nprime + h*ndt][nodes] - curr for h in range(1, N_FWD+1)], axis=1)
        tgt_list.append(((tgt - mu_out) / sd_out).astype(np.float32))

    pred2 = modele(torch.tensor(np.concatenate(X2_list, axis=0)))  
    return criterion(pred2, torch.tensor(np.concatenate(tgt_list, axis=0)))

# ============================================================
# ENTRAÎNEMENT  (loss données + bruit + pushforward)
# ============================================================
historique_train = []
historique_val   = []
historique_pf    = []
meilleure_val    = float("inf")

for epoch in range(1, N_EPOCHS + 1):

    # Montée progressive du poids pushforward (le modèle est nul au début)
    lam_pf = LAMBDA_PF * min(1.0, epoch / PF_WARMUP)

    modele.train()
    perte_train    = 0.0
    perte_pf_total = 0.0

    for X_batch, y_batch in train_loader:
        optimiseur.zero_grad()

        # --- Loss de données : on bruite l'entrée mais la cible reste le delta
        #     PROPRE -> le réseau apprend à bien prédire même sur entrées abîmées. ---

        if NOISE_STD > 0:
            X_in = X_batch + NOISE_STD * torch.randn_like(X_batch)  #We add noise
        else:
            X_in = X_batch

        prediction = modele(X_in)
        data_loss  = criterion(prediction, y_batch)

        # --- Loss pushforward (stabilise le rollout) ---
        if lam_pf > 0:
            pf_loss = pushforward_loss(N_PF_GROUPS)
        else:
            pf_loss = torch.tensor(0.0)

        # --- Loss "au repos" : modele(0) ~= 0, pour ne pas faire dériver les
        #     zones encore immobiles pendant le rollout ---
        rest_loss = criterion(modele(X_rest_n), y_rest_n)

        total_loss = data_loss + lam_pf * pf_loss + LAMBDA_REST * rest_loss
        total_loss.backward()
        optimiseur.step()

        perte_train    += data_loss.item()
        perte_pf_total += pf_loss.item()

    perte_train    /= len(train_loader)
    perte_pf_total /= len(train_loader)

    modele.eval()
    with torch.no_grad():
        pred_val = modele(torch.tensor(X_val)).numpy()
    perte_val = ((pred_val - y_val)**2).mean()
    scheduler.step(perte_val)

    historique_train.append(perte_train)
    historique_val  .append(perte_val)
    historique_pf   .append(perte_pf_total)

    if epoch % 1 == 0:
        print(f"Epoch {epoch:4d}/{N_EPOCHS}  —  "
              f"data: {perte_train:.4f}  |  pushf: {perte_pf_total:.4f}  |  val: {perte_val:.4f}")

    # On garde le meilleur modèle (val minimale)
    if perte_val < meilleure_val:
        meilleure_val = perte_val
        torch.save(modele.state_dict(), SCRIPT_DIR / "model.pth")

# --- GRAPHE : courbe d'apprentissage ---
fig, ax = plt.subplots(figsize=(8, 4))
ax.plot(historique_train, label="Data (train)")
ax.plot(historique_val,   label="Data (val)")
ax.plot(historique_pf,    label="Pushforward")
ax.set_xlabel("Epoch"); ax.set_ylabel("Loss")
ax.set_title("Courbe d'apprentissage (pushforward)")
ax.set_yscale("log"); ax.legend(); ax.grid(True)
plt.tight_layout()
plt.savefig(OUTPUT_DIR / "courbe_apprentissage.png", dpi=150, bbox_inches="tight")
plt.close()

# On recharge le meilleur modèle trouvé pendant l'entraînement
modele.load_state_dict(torch.load(SCRIPT_DIR / "model.pth", weights_only=True))
print(f"Meilleur modèle rechargé — val minimale : {meilleure_val:.6f}")


# ============================================================
# ÉVALUATION TEACHER-FORCING (sur le jeu de test, entrées propres)
# ============================================================
df_test = df[df["split"] == "test"].reset_index(drop=True)

X_new  = df_test[[c + "_n" for c in INPUTS]].values.astype(np.float32)
y_true_n = df_test[OUTPUTS_N].values
y_true = df_test[OUTPUTS].values

modele.eval()
with torch.no_grad():
    y_pred_n = modele(torch.tensor(X_new)).numpy()

# Dénormalisation colonne par colonne
y_pred = np.zeros_like(y_pred_n)
for i, col in enumerate(OUTPUTS):
    y_pred[:, i] = y_pred_n[:, i] * norm_stats.loc[col, "std"] + norm_stats.loc[col, "mean"]

# --- GRAPHE : prédit vs réel (une figure par horizon de sortie) ---
fig, axes = plt.subplots(1, len(OUTPUTS), figsize=(6*len(OUTPUTS), 6), squeeze=False)
axes = axes.flatten()
for i, (ax, col) in enumerate(zip(axes, OUTPUTS)):
    y_r = y_true[:, i]
    y_p = y_pred[:, i]
    ax.scatter(y_r, y_p, alpha=0.4, s=8)
    lim = max(abs(y_r).max(), abs(y_p).max())
    ax.plot([-lim, lim], [-lim, lim], "r--", lw=1, label="prédiction parfaite")
    ax.set_xlabel(f"{col} réel (physique)")
    ax.set_ylabel(f"{col} prédit (physique)")
    
    # --- MODIFICATION ICI : MSE sur les valeurs normalisées ---
    mse_norm = ((y_pred_n[:, i] - y_true_n[:, i]) ** 2).mean()
    r2 = 1 - mse_norm / y_true_n[:, i].var()
    
    ax.set_title(f"{col}\nMSE (norm)={mse_norm:.2e}  |  R²={r2:.3f}")
    ax.legend(); ax.grid(True)

fig.suptitle("Test sur toutes les données test du dataset", fontsize=14)
plt.tight_layout()
plt.savefig(OUTPUT_DIR / "test_predictions.png", dpi=150, bbox_inches="tight")
plt.close()

# Métriques globales dans la console
for i, col in enumerate(OUTPUTS):
    mse_norm = ((y_pred_n[:, i] - y_true_n[:, i]) ** 2).mean()
    r2 = 1 - mse_norm / y_true_n[:, i].var()
    print(f"{col:15s} : MSE (norm) = {mse_norm:.4e}  |  R² = {r2:.4f}")


# ============================================================
# ROLLOUT — paramètres choisis pour rejouer une propagation
# ============================================================
shape, A, omega = SHAPES[0], AMPLITUDES[0], PULSATIONS[0]
x = np.linspace(0, L, Nx)

def u_right(t):
    return u_right_val(shape, A, omega, t)

# Biais "au repos" : sortie du réseau pour une entrée nulle. On le soustrait au
# rollout pour que la zone au repos reste bien à zéro.
# BUG corrigé (même cause que X_rest_n) : x_pos était mis à 0 pour tous les
# nœuds, donc un unique biais_repos (mesuré à x=0) était soustrait partout dans
# le domaine. Le biais réel dépend de x_pos -> un seul par nœud, avec le vrai
# x_pos de ce nœud.
mu = norm_stats.loc[OUTPUTS, "mean"].values
sd = norm_stats.loc[OUTPUTS, "std"].values
Xz = np.tile((-mu_in) / sd_in, (len(nodes), 1)).astype(np.float32)
Xz[:, x_pos_idx] = (x_nodes - mu_in[x_pos_idx]) / sd_in[x_pos_idx]
with torch.no_grad():
    biais_repos = modele(torch.tensor(Xz)).numpy() * sd + mu   # (len(nodes), n_outputs), un biais par nœud

def predict_deltas(U_arr, n, shape, A, omega):
    """État à l'instant n (et son historique dans U_arr) -> deltas dénormalisés
    pour les N_FWD horizons futurs n+ndt, ..., n+N_FWD*ndt."""
    X = build_window([n - lag*ndt for lag in range(M_BACK)], lambda m: U_arr[m])
    with torch.no_grad():
        sortie = modele(torch.tensor(X)).numpy()
    return sortie * sd + mu - biais_repos

def rollout(U_seed, shape, A, omega):
    """Déroule le réseau de manière autorégressive à partir de l'historique
    initial U_seed (vérité pour les history_needed premiers pas) : à chaque
    bloc, on part de l'état courant et on saute de ndt en ndt pour prédire
    les N_FWD instants suivants, jusqu'à couvrir tout l'horizon Nt."""
    U_ = U_seed.copy()
    for n in range(history_needed, Nt - N_FWD * ndt + 1, N_FWD*ndt):
        deltas = predict_deltas(U_, n, shape, A, omega)
        for h in range(1, N_FWD + 1):
            s = n + h*ndt
            u_new = np.clip(U_[n, nodes] + deltas[:, h-1], -CLAMP_AMP, CLAMP_AMP)
            U_[s, nodes]     = u_new
            U_[s, :i_left+1] = 0.0
            U_[s, i_right:]  = u_right_val(shape, A, omega, s * dt)

            # Lissage Laplacien léger : casse le mode "damier" haute fréquence
            # que u_xx amplifierait au bloc suivant, sans amortir l'onde (basse
            # fréquence).
            if SMOOTH_ALPHA > 0:
                j0, j1 = i_left + 1, i_right        # nœuds intérieurs (hors bords)
                lap = U_[s, j0-1:j1-1] - 2*U_[s, j0:j1] + U_[s, j0+1:j1+1]
                U_[s, j0:j1] += SMOOTH_ALPHA * lap
    return U_

def chrono(fonction, n_repeat=15, n_warmup=3):
    for _ in range(n_warmup):
        fonction()
    t = np.array([(lambda: (time.perf_counter(), fonction(), time.perf_counter())[::2])() for _ in range(n_repeat)])
    d = t[:, 1] - t[:, 0]
    return d.mean(), d.std(), np.median(d)

# ----------------------------------------------------------------
# 1) VRAIE simulation (référence)
# ----------------------------------------------------------------
t0 = time.perf_counter()

U_reel = np.zeros((Nt + 1, Ntot))
u, u_1 = np.zeros(Ntot), np.zeros(Ntot)

for n in range(Nt):
    u_new = np.zeros(Ntot)
    u_new[i_left:i_right+1] = (2*u[i_left:i_right+1] - u_1[i_left:i_right+1] + CFL**2 * (u[i_left-1:i_right] - 2*u[i_left:i_right+1] + u[i_left+1:i_right+2]))
    u_new[:i_left+1] = 0.0
    u_new[i_right:]  = u_right((n+1)*dt)
    u_1, u = u.copy(), u_new
    U_reel[n + 1] = u.copy()

time_phys = time.perf_counter() - t0

# u_tt et u_xx réels (pour le graphe de vérification de l'EDP)
ureel_tt_dict = np.zeros((Nt, Ntot))
ureel_xx_dict = np.zeros((Nt, Ntot))
for n in range(1, Nt):
    u_prev, u_curr, u_next = U_reel[n - 1], U_reel[n], U_reel[n + 1]
    ureel_tt_dict[n, i_left:i_right+1] = (u_next[i_left:i_right+1] - 2*u_curr[i_left:i_right+1] + u_prev[i_left:i_right+1]) / dt**2
    ureel_xx_dict[n, i_left:i_right+1] = (u_curr[i_left-1:i_right] - 2*u_curr[i_left:i_right+1] + u_curr[i_left+1:i_right+2]) / dx**2

# --- GRAPHE : u_tt en fonction de u_xx (réel) ---
plt.figure()
for n in [50, 100, 200]:
    plt.scatter(ureel_xx_dict[n, i_left+1:i_right], ureel_tt_dict[n, i_left+1:i_right], s=10, label=f"n = {n}")
plt.xlabel("u_xx (réel)"); plt.ylabel("u_tt (réel)")
plt.legend(); plt.grid(); plt.xlim(-10, 10); plt.ylim(-10, 10)
plt.title("u_tt en fonction de u_xx (real)")
plt.savefig(OUTPUT_DIR / "utt_uxx_reel.png", dpi=150, bbox_inches="tight")
plt.close()

# ----------------------------------------------------------------
# 2) Simulation PRÉDITE (rollout du réseau sur ses propres sorties)
# ----------------------------------------------------------------
t0 = time.perf_counter()

history_needed = M_BACK * ndt
U_seed = np.zeros((Nt + 1, Ntot))
for m in range(history_needed + 1):      # on amorce avec la vérité
    U_seed[m] = U_reel[m]

U = rollout(U_seed, shape, A, omega)

time_pred = time.perf_counter() - t0

# u_tt et u_xx prédits (pour le graphe de vérification de l'EDP)
upred_tt_dict = np.zeros((Nt, Ntot))
upred_xx_dict = np.zeros((Nt, Ntot))
for n in range(ndt, Nt - ndt + 1, ndt):
    u_prev, u_curr, u_next = U[n - ndt], U[n], U[n + ndt]
    upred_tt_dict[n, i_left:i_right+1] = (u_next[i_left:i_right+1] - 2*u_curr[i_left:i_right+1] + u_prev[i_left:i_right+1]) / (ndt*dt)**2
    upred_xx_dict[n, i_left:i_right+1] = (u_curr[i_left-1:i_right] - 2*u_curr[i_left:i_right+1] + u_curr[i_left+1:i_right+2]) / dx**2

# --- GRAPHE : u_tt en fonction de u_xx (prédit) ---
plt.figure()
for n in [5, 100, 150]:
    plt.scatter(upred_xx_dict[n, i_left+1:i_right], upred_tt_dict[n, i_left+1:i_right], s=10, label=f"n = {n}")
plt.xlabel("u_xx (predit)"); plt.ylabel("u_tt (predit)")
plt.grid(); plt.xlim(-10, 10); plt.ylim(-10, 10); plt.legend()
plt.title("u_tt en fonction de u_xx (prediction)")
plt.savefig(OUTPUT_DIR / "utt_uxx_predit.png", dpi=150, bbox_inches="tight")
plt.close()

print("temps physique :", round(time_phys, 6))
print("temps predit   :", round(time_pred, 6))

# ============================================================
# GRAPHES FINAUX DU ROLLOUT
# ============================================================
def l2_rel(pred, true, eps=1e-12):
    return np.linalg.norm(pred - true) / (np.linalg.norm(true) + eps)

def smape(pred, true):
    m = true != 0
    return np.mean(2*np.abs(true[m] - pred[m]) / (np.abs(true[m]) + np.abs(pred[m])))

# --- GRAPHE (animation) : propagation de l'onde réelle vs prédite + erreur ---
# U n'est rempli qu'aux multiples de ndt pendant le rollout -> on anime seulement
# ces pas, sinon la courbe prédite retomberait à zéro.
frames = np.arange(0, Nt + 1, ndt)

fig_anim, (axA, axB) = plt.subplots(2, 1, figsize=(9, 7), sharex=True)

ligne_reel, = axA.plot([], [], "r",   lw=2, label="réel")
ligne_pred, = axA.plot([], [], "b--", lw=2, label="prédit")
ymax = np.abs(U_reel[:, nodes]).max() * 1.2
axA.set_xlim(0, L); axA.set_ylim(-ymax, ymax)
axA.set_ylabel("u"); axA.legend(loc="upper right"); axA.grid(True)

ligne_err, = axB.plot([], [], "k", lw=1.5, label="|prédit - réel|")
err_max = max(np.max([np.abs(U[m, nodes] - U_reel[m, nodes]).max() for m in frames]) * 1.2, 1e-9)
axB.set_xlim(0, L); axB.set_ylim(0, err_max)
axB.set_xlabel("x"); axB.set_ylabel("erreur absolue"); axB.legend(loc="upper right"); axB.grid(True)

titre = fig_anim.suptitle("")

def maj(m):
    ligne_reel.set_data(x, U_reel[m, nodes])
    ligne_pred.set_data(x, U[m, nodes])
    ligne_err.set_data(x, np.abs(U[m, nodes] - U_reel[m, nodes]))
    titre.set_text(f"Propagation de l'onde — t = {m*dt:.3f}  (pas {m})")
    return ligne_reel, ligne_pred, ligne_err, titre

anim = animation.FuncAnimation(fig_anim, maj, frames=frames, interval=50, blit=False)
anim.save(OUTPUT_DIR / "propagation_onde.gif", writer="pillow", fps=20, dpi=110)
plt.close(fig_anim)
print(f"Animation sauvegardée : {OUTPUT_DIR / 'propagation_onde.gif'}")

# --- GRAPHE : erreur (L2 relative et Linf) en fonction du temps ---
steps  = np.arange(2*ndt, Nt + 1, ndt)          # indices réellement remplis
t_axis = steps * dt
l2_list   = [l2_rel(U[k, nodes], U_reel[k, nodes])         for k in steps]
linf_list = [np.max(np.abs(U[k, nodes] - U_reel[k, nodes])) for k in steps]

plt.figure(figsize=(9, 5))
plt.plot(t_axis, l2_list,   "o-", ms=3, label="erreur L2 relative")
plt.plot(t_axis, linf_list, "s-", ms=3, label="erreur max absolue (Linf)")
plt.yscale("log")
plt.xlabel("t"); plt.ylabel("erreur"); plt.grid(True, which="both"); plt.legend()
plt.title("Erreur du rollout en fonction du temps")
plt.savefig(OUTPUT_DIR / "erreur_temps.png", dpi=150, bbox_inches="tight")
plt.close()

import openpyxl
xlsx_path = SCRIPT_DIR.parent / "Comparaisons" / "comparative_table.xlsx"
erreur = l2_list
if not xlsx_path.exists():
    print(f"[info] Tableau comparatif absent ({xlsx_path}) — export ignoré.")
else:
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        col_uxx = None
        for cell in ws[1]:
            if cell.value == "Ut_Uxx_PF_Lap":
                col_uxx = cell.column
                break
        if col_uxx is None:
            raise ValueError("Colonne 'Ut_Uxx_PF_Lap' introuvable dans la première ligne.")
        col_time = col_uxx - 1
        for i, (t, e) in enumerate(zip(t_axis, erreur), start=2):
            ws.cell(row=i, column=col_time, value=float(t))
            ws.cell(row=i, column=col_uxx,  value=float(e))
        wb.save(xlsx_path)
        print(f"{len(erreur)} valeurs écrites dans {xlsx_path} (colonnes time / Ut_Uxx_PF_Lap).")
    except Exception as exc:
        print(f"[avertissement] Échec de l'export vers {xlsx_path} : {exc}")

# --- GRAPHE : sMAPE en fonction du temps ---
smape_list = [100.0 * smape(U[k, nodes], U_reel[k, nodes]) for k in steps]

plt.figure(figsize=(9, 5))
plt.plot(t_axis, smape_list, "s-", ms=3, label="sMAPE")
plt.xlabel("t"); plt.ylabel("erreur (%)"); plt.grid(True); plt.legend()
plt.title("sMAPE du rollout en fonction du temps")
plt.savefig(OUTPUT_DIR / "smape_temps.png", dpi=150, bbox_inches="tight")
plt.close()

import torch
torch.set_num_threads(1)   # comparaison reproductible

def fd_once():
    U_ = np.zeros((Nt + 1, Ntot)); u, u_1 = np.zeros(Ntot), np.zeros(Ntot)
    for k in range(Nt):
        un = np.zeros(Ntot)
        un[i_left:i_right+1] = (2*u[i_left:i_right+1] - u_1[i_left:i_right+1] + CFL**2*(u[i_left-1:i_right]-2*u[i_left:i_right+1]+u[i_left+1:i_right+2]))
        un[:i_left+1] = 0.0; un[i_right:] = u_right_val(shape, A, omega, (k+1)*dt)
        u_1, u = u.copy(), un; U_[k+1] = u.copy()
    return U_

def rollout_once():
    return rollout(U_seed, shape, A, omega)

_, _, med_fd = chrono(fd_once)
_, s_nn, med_nn = chrono(rollout_once)
print(f"FD (réel)   : {med_fd*1e3:7.3f} ms")
print(f"NN (rollout): {med_nn*1e3:7.3f} ms  (±{s_nn*1e3:.3f})")
print(f"speedup FD/NN = {med_fd/med_nn:.2f}x   (>1 = le réseau est plus rapide)")

from torch.utils.flop_counter import FlopCounterMode
with FlopCounterMode(display=False) as fc:
    modele(torch.zeros((len(nodes), len(INPUTS))))
n_appels = len(range(history_needed, Nt - N_FWD*ndt + 1, N_FWD*ndt))
print(f"FLOPs réseau (rollout) ≈ {fc.get_total_flops()*n_appels:,.0f}")
print(f"FLOPs FD (~Nt*Nx*8)    ≈ {Nt*Nx*8:,}")

# =====================================================
# RESUME : valeurs phares -> outputs/resume.txt
# =====================================================
# Erreurs finales (dernier pas rempli) ET max sur tout le rollout
l2_final,    l2_max    = l2_list[-1],    max(l2_list)
linf_final,  linf_max  = linf_list[-1],  max(linf_list)
smape_final, smape_max = smape_list[-1], max(smape_list)

with open(OUTPUT_DIR / "resume.txt", "w") as f:
    f.write("=====  RESUME DU RUN  =====\n\n")

    f.write("--- Configuration ---\n")
    f.write(f"Grille          : Nt={Nt}, Nx={Nx}, SS={SS}, ndt={ndt}\n")
    f.write(f"Rollout (shape,A,w) : shape={shape}, A={A}, omega={omega}\n")
    f.write(f"Formes d'onde   : {SHAPES}\n")
    f.write(f"Dataset         : {len(df):,} lignes\n")
    f.write(f"Features        : {len(INPUTS)} entrees, {len(OUTPUTS)} sortie(s)\n")
    for s in ["train", "val", "test"]:
        n = (df["split"] == s).sum()
        f.write(f"  split {s:5s}   : {n:>8,} lignes ({100*n/len(df):.1f} %)\n")
    f.write(f"Parametres NN   : {n_params:,}\n\n")

    f.write("--- Entrainement ---\n")
    f.write(f"Val minimale    : {meilleure_val:.6e}\n")
    for i, col in enumerate(OUTPUTS):
        mse_norm = ((y_pred_n[:, i] - y_true_n[:, i]) ** 2).mean()
        r2 = 1 - mse_norm / y_true_n[:, i].var()
        f.write(f"{col:15s} : MSE (norm) = {mse_norm:.4e} | R2 = {r2:.4f}\n")
    f.write("\n")

    f.write("--- Temps d'execution ---\n")
    f.write(f"Simulation reelle (FD) : {time_phys:.6f} s\n")
    f.write(f"Rollout predit (NN)    : {time_pred:.6f} s\n")
    f.write(f"Ratio NN / FD          : {time_pred/time_phys:.2f}\n\n")

    f.write("--- Erreurs du rollout ---\n")
    f.write(f"L2 relative  : finale = {l2_final:.4e}  |  max = {l2_max:.4e}\n")
    f.write(f"Linf absolue : finale = {linf_final:.4e}  |  max = {linf_max:.4e}\n")
    f.write(f"sMAPE (%)    : finale = {smape_final:.3f}  |  max = {smape_max:.3f}\n")

print(f"Resume sauvegarde : {OUTPUT_DIR / 'resume.txt'}")